import json
from openpyxl import Workbook
from openpyxl import load_workbook

#workbook = Workbook()
wb_total = load_workbook(filename = 'total_count.xlsx')
sheet_total = wb_total.active

wb_1 = load_workbook(filename = 'count_path_1.xlsx')
sheet_1 = wb_1.active


wb_2 = load_workbook(filename = 'count_path_2.xlsx')
sheet_2 = wb_2.active


wb_3 = load_workbook(filename = 'count_path_3.xlsx')
sheet_3 = wb_3.active


wb_4 = load_workbook(filename = 'count_path_4.xlsx')
sheet_4 = wb_4.active

wb_5 = load_workbook(filename = 'count_path_5.xlsx')
sheet_5 = wb_5.active

  
# Opening JSON file
f = open('summary_falcon_7M_1.json') #only opens one exploration


exploredNodes = []

#mousemove
count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0



  
# returns JSON object as 
# a dictionary
data = json.load(f)

for i in data["0"]:
    
    ### mousemove
    
    if(i == "mousemove on #distance canvas.marks , transition: 0--->0"):
        count_mousemove_distance = count_mousemove_distance + 1
    
    if(i == "mousemove on #arrival canvas.marks , transition: 0--->0"):
        count_mousemove_arrival = count_mousemove_arrival + 1
    
    if(i == "mousemove on #count canvas.marks , transition: 0--->0"):
        count_mousemove_count = count_mousemove_count + 1
    
    if(i == "mousemove on #departure canvas.marks , transition: 0--->0"):
        count_mousemove_departure = count_mousemove_departure + 1
    
    if(i == "mousemove on #airtime canvas.marks , transition: 0--->0"):
        count_mousemove_airtime = count_mousemove_airtime + 1
    
    if(i == "mousemove on #delay canvas.marks , transition: 0--->0"):
        count_mousemove_delay = count_mousemove_delay + 1
        
        ### mousemove brushing

    if(i == "mousemove on departure, transition: 6--->6"):
        count_mousemove_brush_departure +=1
    
    if(i == "mousemove on arrival, transition: 5--->5"):
        count_mousemove_brush_arrival +=1
    
    if(i == "mousemove on airtime, transition: 7--->7"):
        count_mousemove_brush_airtime +=1
    
    if(i == "mousemove on delay, transition: 8--->8"):
        count_mousemove_brush_delay +=1
    
    if(i == "mousemove on distance, transition: 4--->4"):
        count_mousemove_brush_distance +=1

    ### mousemove count

    if(i == "mousemove on count, transition: 1--->1"):
        special_mousemove_count+=1

    ### mousedown

    if(i == "mousedown on distance, transition: 0--->4"):
        count_mousedown_distance+=1

    if(i == "mousedown on departure, transition: 0--->6"):
        count_mousedown_departure+=1

    if(i == "mousedown on arrival, transition: 0--->5"):
        count_mousedown_arrival+=1

    if(i == "mousedown on airtime, transition: 0--->7"):
        count_mousedown_airtime+=1

    if(i == "mousedown on delay, transition: 0--->8"):
        count_mousedown_delay +=1

    if(i == "mousedown on count, transition: 0--->1"):
        count_mousedown_count +=1
    
    ### mouseup

    if(i == "mouseup on distance, transition: 4--->0"):
        count_mouseup_distance+=1

    if(i == "mouseup on departure, transition: 6--->0"):
        count_mouseup_departure+=1

    if(i == "mouseup on arrival, transition: 5--->0"):
        count_mouseup_arrival+=1

    if(i == "mouseup on airtime, transition: 7--->0"):
        count_mouseup_airtime+=1

    if(i == "mouseup on delay, transition: 8--->0"):
        count_mouseup_delay +=1

    ### click

    if(i == "click on #departure canvas.marks , transition: 0--->0"):
        count_click_departure+=1

    if(i == "click on #distance canvas.marks , transition: 0--->0"):
        count_click_distance+=1

    if(i == "click on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "click on #arrival canvas.marks , transition: 0--->0"):
        count_click_arrival+=1

    if(i == "click on #delay canvas.marks , transition: 0--->0"):
        count_click_delay +=1

    if(i == "click on #count canvas.marks , transition: 0--->0"):
        count_click_count +=1

    if(i == "click on #count canvas.marks , transition: 1--->1"):
        special_click_count +=1

    ### dblclick


    if(i == "dblclick on #departure canvas.marks , transition: 0--->0"):
        count_dblclick_departure+=1

    if(i == "dblclick on #distance canvas.marks , transition: 0--->0"):
        count_dblclick_distance+=1

    if(i == "dblclick on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "dblclick on #arrival canvas.marks , transition: 0--->0"):
        count_dblclick_arrival+=1

    if(i == "dblclick on #delay canvas.marks , transition: 0--->0"):
        count_dblclick_delay +=1

    if(i == "dblclick on #count canvas.marks , transition: 0--->0"):
        count_dblclick_count +=1

    ### wheel

    if(i == "wheel on #distance canvas.marks , transition: 0--->0"):
        count_wheel_distance +=1
    
    if(i == "wheel on #departure canvas.marks , transition: 0--->0"):
        count_wheel_departure +=1
    
    if(i == "wheel on #airtime canvas.marks , transition: 0--->0"):
        count_wheel_airtime +=1
    
    if(i == "wheel on #arrival canvas.marks , transition: 0--->0"):
        count_wheel_arrival +=1
    
    if(i == "wheel on #delay canvas.marks , transition: 0--->0"):
        count_wheel_delay +=1

    ### special wheel

    if(i == "wheel on departure while brushing, transition: 6--->6"):
        special_wheel_departure +=1
    
    if(i == "wheel on arrival while brushing, transition: 5--->5"):
        special_wheel_arrival +=1
    
    if(i == "wheel on airtime while brushing, transition: 7--->7"):
        special_wheel_airtime +=1
    
    if(i == "wheel on delay while brushing, transition: 8--->8"):
        special_wheel_delay +=1
    
    if(i == "wheel on distance while brushing, transition: 4--->4"):
        special_wheel_distance +=1

    ### mouseout

    if(i == "mouseout on #distance canvas.marks , transition: 0--->0"):
        count_mouseout_distance +=1
    
    if(i == "mouseout on #departure canvas.marks , transition: 0--->0"):
        count_mouseout_departure +=1
    
    if(i == "mouseout on #airtime canvas.marks , transition: 0--->0"):
        count_mouseout_airtime +=1
    
    if(i == "mouseout on #arrival canvas.marks , transition: 0--->0"):
        count_mouseout_arrival +=1
    
    if(i == "mouseout on #delay canvas.marks , transition: 0--->0"):
        count_mouseout_delay +=1
    
    if(i == "mouseout on #count canvas.marks , transition: 0--->0"):
        count_mouseout_count +=1

    if(i == "mouseout on #count canvas.marks , transition: 1--->1"):
        special_mouseout_count +=1

    ### special mouseout

    if(i == "mouseout on departure while brushing, transition: 6--->6"):
        special_mouseout_departure +=1
    
    if(i == "mouseout on arrival while brushing, transition: 5--->5"):
        special_mouseout_arrival +=1
    
    if(i == "mouseout on airtime while brushing, transition: 7--->7"):
        special_mouseout_airtime +=1
    
    if(i == "mouseout on delay while brushing, transition: 8--->8"):
        special_mouseout_delay +=1
    
    if(i == "mouseout on distance while brushing, transition: 4--->4"):
        special_mouseout_distance +=1



exploredNodes.append("----------------------PATH 1------------------")

if(count_mousemove_departure) != 0:exploredNodes.append("number of mousemove on departure, transition: 0--->0 = " + str(count_mousemove_departure))
if(count_mousemove_delay) != 0:exploredNodes.append("number of mousemove on delay, transition: 0--->0 = " + str(count_mousemove_delay))
if(count_mousemove_airtime) != 0:exploredNodes.append("number of mousemove on airtime, transition: 0--->0 = " + str(count_mousemove_airtime))
if(count_mousemove_arrival) != 0:exploredNodes.append("number of mousemove on arrival, transition: 0--->0 = " + str(count_mousemove_arrival))
if(count_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 0--->0 = " + str(count_mousemove_count))
if(count_mousemove_distance) != 0:exploredNodes.append("number of mousemove on distance, transition: 0--->0 = " + str(count_mousemove_distance))

if(count_mousemove_brush_departure) != 0:exploredNodes.append("number of special mousemove on departure, transition: 6--->6 = " + str(count_mousemove_brush_departure))
if(count_mousemove_brush_arrival) != 0:exploredNodes.append("number of special mousemove on arrival, transition: 5--->5 = " + str(count_mousemove_brush_arrival))
if(count_mousemove_brush_airtime) != 0:exploredNodes.append("number of special mousemove on airtime, transition: 7--->7 = " + str(count_mousemove_brush_airtime))
if(count_mousemove_brush_delay) != 0:exploredNodes.append("number of special mousemove on delay, transition: 8--->8 = " + str(count_mousemove_brush_delay))
if(count_mousemove_brush_distance) != 0:exploredNodes.append("number of special mousemove on distance, transition: 4--->4 = " + str(count_mousemove_brush_distance))

if(special_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 1--->1 = " + str(special_mousemove_count))

if(count_mousedown_distance) != 0:exploredNodes.append("number of mousedown on distance, transition: 0--->4 = " + str(count_mousedown_distance))
if(count_mousedown_departure) != 0:exploredNodes.append("number of mousedown on departure, transition: 0--->6 = " + str(count_mousedown_departure))
if(count_mousedown_arrival) != 0:exploredNodes.append("number of mousedown on arrival, transition: 0--->5 = " + str(count_mousedown_arrival))
if(count_mousedown_airtime) != 0:exploredNodes.append("number of mousedown on airtime, transition: 0--->7 = " + str(count_mousedown_airtime))
if(count_mousedown_delay) != 0:exploredNodes.append("number of mousedown on delay, transition: 0--->8 = " + str(count_mousedown_delay))
if(count_mousedown_count) != 0:exploredNodes.append("number of mousedown on count, transition: 0--->1 = " + str(count_mousedown_count))

if(count_mouseup_distance) != 0:exploredNodes.append("number of mouseup on distance, transition: 4--->0 = " + str(count_mouseup_distance))
if(count_mouseup_departure) != 0:exploredNodes.append("number of mouseup on departure, transition: 6--->0 = " + str(count_mouseup_departure))
if(count_mouseup_arrival) != 0:exploredNodes.append("number of mouseup on arrival, transition: 5--->0 = " + str(count_mouseup_arrival))
if(count_mouseup_airtime) != 0:exploredNodes.append("number of mouseup on airtime, transition: 7--->0 = " + str(count_mouseup_airtime))
if(count_mouseup_delay) != 0:exploredNodes.append("number of mouseup on delay, transition: 8--->0 = " + str(count_mouseup_delay))

if(count_click_distance) != 0:exploredNodes.append("number of click on distance, transition: 0--->0 = " + str(count_click_distance))
if(count_click_departure) != 0:exploredNodes.append("number of click on departure, transition: 0--->0 = " + str(count_click_departure))
if(count_click_arrival) != 0:exploredNodes.append("number of click on arrival, transition: 0--->0 = " + str(count_click_arrival))
if(count_click_airtime) != 0:exploredNodes.append("number of click on airtime, transition: 0--->0 = " + str(count_click_airtime))
if(count_click_delay) != 0:exploredNodes.append("number of click on delay, transition: 0--->0 = " + str(count_click_delay))
if(count_click_count) != 0:exploredNodes.append("number of click on count, transition: 0--->0 = " + str(count_click_count))
if(special_click_count) != 0:exploredNodes.append("number of special click on count, transition: 1--->1 = " + str(special_click_count))

if(count_dblclick_distance) != 0:exploredNodes.append("number of dblclick on distance, transition: 0--->0 = " + str(count_dblclick_distance))
if(count_dblclick_departure) != 0:exploredNodes.append("number of dblclick on departure, transition: 0--->0 = " + str(count_dblclick_departure))
if(count_dblclick_arrival) != 0:exploredNodes.append("number of dblclick on arrival, transition: 0--->0 = " + str(count_dblclick_arrival))
if(count_dblclick_airtime) != 0:exploredNodes.append("number of dblclick on airtime, transition: 0--->0 = " + str(count_dblclick_airtime))
if(count_dblclick_delay) != 0:exploredNodes.append("number of dblclick on delay, transition: 0--->0 = " + str(count_dblclick_delay))
if(count_dblclick_count) != 0:exploredNodes.append("number of dblclick on count, transition: 0--->0 = " + str(count_dblclick_count))

if(count_wheel_distance) != 0:exploredNodes.append("number of wheel on distance, transition: 0--->0 = " + str(count_wheel_distance))
if(count_wheel_departure) != 0:exploredNodes.append("number of wheel on departure, transition: 0--->0 = " + str(count_wheel_departure))
if(count_wheel_arrival) != 0:exploredNodes.append("number of wheel on arrival, transition: 0---> = " + str(count_wheel_arrival))
if(count_wheel_airtime) != 0:exploredNodes.append("number of wheel on airtime, transition: 0--->0 = " + str(count_wheel_airtime))
if(count_wheel_delay) != 0:exploredNodes.append("number of wheel on delay, transition: 0--->0 = " + str(count_wheel_delay))

if(special_wheel_distance) != 0:exploredNodes.append("number of special wheel on distance, transition: 4--->4 = " + str(special_wheel_distance))
if(special_wheel_departure) != 0:exploredNodes.append("number of special wheel on departure, transition: 6--->6 = " + str(special_wheel_departure))
if(special_wheel_arrival) != 0:exploredNodes.append("number of special wheel on arrival, transition: 5--->5 = " + str(special_wheel_arrival))
if(special_wheel_airtime) != 0:exploredNodes.append("number of special wheel on airtime, transition: 7--->7 = " + str(special_wheel_airtime))
if(special_wheel_delay) != 0:exploredNodes.append("number of special wheel on delay, transition: 8--->8 = " + str(special_wheel_delay))

if(count_mouseout_distance) != 0:exploredNodes.append("number of mouseout on distance, transition: 0--->0 = " + str(count_mouseout_distance))
if(count_mouseout_departure) != 0:exploredNodes.append("number of mouseout on departure, transition: 0--->0 = " + str(count_mouseout_departure))
if(count_mouseout_arrival) != 0:exploredNodes.append("number of mouseout on arrival, transition: 0--->0 = " + str(count_mouseout_arrival))
if(count_mouseout_airtime) != 0:exploredNodes.append("number of mouseout on airtime, transition: 0--->0 = " + str(count_mouseout_airtime))
if(count_mouseout_delay) != 0:exploredNodes.append("number of mouseout on delay, transition: 0--->0 = " + str(count_mouseout_delay))

if(special_mouseout_distance) != 0:exploredNodes.append("number of special mouseout on distance, transition: 4--->4 = " + str(special_mouseout_distance))
if(special_mouseout_departure) != 0:exploredNodes.append("number of special mouseout on departure, transition: 6--->6 = " + str(special_mouseout_departure))
if(special_mouseout_arrival) != 0:exploredNodes.append("number of special mouseout on arrival, transition: 5--->5 = " + str(special_mouseout_arrival))
if(special_mouseout_airtime) != 0:exploredNodes.append("number of special mouseout on airtime, transition: 7--->7 = " + str(special_mouseout_airtime))
if(special_mouseout_delay) != 0:exploredNodes.append("number of special mouseout on delay, transition: 8--->8 = " + str(special_mouseout_delay))
if(special_mouseout_count) != 0:exploredNodes.append("number of special mouseout on count, transition: 1--->1 = " + str(special_mouseout_count))

#save on excel file

sheet_total["A3"].value += count_mousemove_departure
sheet_total["B3"].value += count_mousemove_distance
sheet_total["C3"].value += count_mousemove_delay
sheet_total["D3"].value += count_mousemove_airtime
sheet_total["E3"].value += count_mousemove_arrival
sheet_total["F3"].value += count_mousemove_count
sheet_total["G3"].value += special_mousemove_count

sheet_total["A6"].value += count_mousedown_departure
sheet_total["B6"].value += count_mousedown_distance
sheet_total["C6"].value += count_mousedown_delay
sheet_total["D6"].value += count_mousedown_airtime
sheet_total["E6"].value += count_mousedown_arrival
sheet_total["F6"].value += count_mousedown_count

sheet_total["A10"].value += count_mousemove_brush_departure
sheet_total["B10"].value += count_mousemove_brush_distance
sheet_total["C10"].value += count_mousemove_brush_delay
sheet_total["D10"].value += count_mousemove_brush_airtime
sheet_total["E10"].value += count_mousemove_brush_arrival

sheet_total["A13"].value += count_mouseup_departure
sheet_total["B13"].value += count_mouseup_distance
sheet_total["C13"].value += count_mouseup_delay
sheet_total["D13"].value += count_mouseup_airtime
sheet_total["E13"].value += count_mouseup_arrival

sheet_total["A17"].value += count_click_departure
sheet_total["B17"].value += count_click_distance
sheet_total["C17"].value += count_click_delay
sheet_total["D17"].value += count_click_airtime
sheet_total["E17"].value += count_click_arrival
sheet_total["F17"].value += count_click_count
sheet_total["G17"].value += special_click_count

sheet_total["A21"].value += count_mouseout_departure
sheet_total["B21"].value += count_mouseout_distance
sheet_total["C21"].value += count_mouseout_delay
sheet_total["D21"].value += count_mouseout_airtime
sheet_total["E21"].value += count_mouseout_arrival
sheet_total["F21"].value += count_mouseout_count

sheet_total["A25"].value += special_mouseout_departure
sheet_total["B25"].value += special_mouseout_distance
sheet_total["C25"].value += special_mouseout_delay
sheet_total["D25"].value += special_mouseout_airtime
sheet_total["E25"].value += special_mouseout_arrival
sheet_total["F25"].value += special_mouseout_count

sheet_total["A29"].value += count_wheel_departure
sheet_total["B29"].value += count_wheel_distance
sheet_total["C29"].value += count_wheel_delay
sheet_total["D29"].value += count_wheel_airtime
sheet_total["E29"].value += count_wheel_arrival


sheet_total["A33"].value += special_wheel_departure
sheet_total["B33"].value += special_wheel_distance
sheet_total["C33"].value += special_wheel_delay
sheet_total["D33"].value += special_wheel_airtime
sheet_total["E33"].value += special_wheel_arrival

sheet_total["A37"].value += count_dblclick_departure
sheet_total["B37"].value += count_dblclick_distance
sheet_total["C37"].value += count_dblclick_delay
sheet_total["D37"].value += count_dblclick_airtime
sheet_total["E37"].value += count_dblclick_arrival
sheet_total["F37"].value += count_dblclick_count

wb_total.save(filename="total_count.xlsx")

#saving to the relative path excel file

sheet_1["A3"].value += count_mousemove_departure
sheet_1["B3"].value += count_mousemove_distance
sheet_1["C3"].value += count_mousemove_delay
sheet_1["D3"].value += count_mousemove_airtime
sheet_1["E3"].value += count_mousemove_arrival
sheet_1["F3"].value += count_mousemove_count
sheet_1["G3"].value += special_mousemove_count

sheet_1["A6"].value += count_mousedown_departure
sheet_1["B6"].value += count_mousedown_distance
sheet_1["C6"].value += count_mousedown_delay
sheet_1["D6"].value += count_mousedown_airtime
sheet_1["E6"].value += count_mousedown_arrival
sheet_1["F6"].value += count_mousedown_count

sheet_1["A10"].value += count_mousemove_brush_departure
sheet_1["B10"].value += count_mousemove_brush_distance
sheet_1["C10"].value += count_mousemove_brush_delay
sheet_1["D10"].value += count_mousemove_brush_airtime
sheet_1["E10"].value += count_mousemove_brush_arrival

sheet_1["A13"].value += count_mouseup_departure
sheet_1["B13"].value += count_mouseup_distance
sheet_1["C13"].value += count_mouseup_delay
sheet_1["D13"].value += count_mouseup_airtime
sheet_1["E13"].value += count_mouseup_arrival

sheet_1["A17"].value += count_click_departure
sheet_1["B17"].value += count_click_distance
sheet_1["C17"].value += count_click_delay
sheet_1["D17"].value += count_click_airtime
sheet_1["E17"].value += count_click_arrival
sheet_1["F17"].value += count_click_count
sheet_1["G17"].value += special_click_count

sheet_1["A21"].value += count_mouseout_departure
sheet_1["B21"].value += count_mouseout_distance
sheet_1["C21"].value += count_mouseout_delay
sheet_1["D21"].value += count_mouseout_airtime
sheet_1["E21"].value += count_mouseout_arrival
sheet_1["F21"].value += count_mouseout_count

sheet_1["A25"].value += special_mouseout_departure
sheet_1["B25"].value += special_mouseout_distance
sheet_1["C25"].value += special_mouseout_delay
sheet_1["D25"].value += special_mouseout_airtime
sheet_1["E25"].value += special_mouseout_arrival
sheet_1["F25"].value += special_mouseout_count

sheet_1["A29"].value += count_wheel_departure
sheet_1["B29"].value += count_wheel_distance
sheet_1["C29"].value += count_wheel_delay
sheet_1["D29"].value += count_wheel_airtime
sheet_1["E29"].value += count_wheel_arrival


sheet_1["A33"].value += special_wheel_departure
sheet_1["B33"].value += special_wheel_distance
sheet_1["C33"].value += special_wheel_delay
sheet_1["D33"].value += special_wheel_airtime
sheet_1["E33"].value += special_wheel_arrival

sheet_1["A37"].value += count_dblclick_departure
sheet_1["B37"].value += count_dblclick_distance
sheet_1["C37"].value += count_dblclick_delay
sheet_1["D37"].value += count_dblclick_airtime
sheet_1["E37"].value += count_dblclick_arrival
sheet_1["F37"].value += count_dblclick_count

wb_1.save(filename="count_path_1.xlsx")

#setting back to 0 the variables

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0


for i in data["1"]:
    ### mousemove
    
    if(i == "mousemove on #distance canvas.marks , transition: 0--->0"):
        count_mousemove_distance = count_mousemove_distance + 1
    
    if(i == "mousemove on #arrival canvas.marks , transition: 0--->0"):
        count_mousemove_arrival = count_mousemove_arrival + 1
    
    if(i == "mousemove on #count canvas.marks , transition: 0--->0"):
        count_mousemove_count = count_mousemove_count + 1
    
    if(i == "mousemove on #departure canvas.marks , transition: 0--->0"):
        count_mousemove_departure = count_mousemove_departure + 1
    
    if(i == "mousemove on #airtime canvas.marks , transition: 0--->0"):
        count_mousemove_airtime = count_mousemove_airtime + 1
    
    if(i == "mousemove on #delay canvas.marks , transition: 0--->0"):
        count_mousemove_delay = count_mousemove_delay + 1
        
        ### mousemove brushing

    if(i == "mousemove on departure, transition: 6--->6"):
        count_mousemove_brush_departure +=1
    
    if(i == "mousemove on arrival, transition: 5--->5"):
        count_mousemove_brush_arrival +=1
    
    if(i == "mousemove on airtime, transition: 7--->7"):
        count_mousemove_brush_airtime +=1
    
    if(i == "mousemove on delay, transition: 8--->8"):
        count_mousemove_brush_delay +=1
    
    if(i == "mousemove on distance, transition: 4--->4"):
        count_mousemove_brush_distance +=1

    ### mousemove count

    if(i == "mousemove on count, transition: 1--->1"):
        special_mousemove_count+=1

    ### mousedown

    if(i == "mousedown on distance, transition: 0--->4"):
        count_mousedown_distance+=1

    if(i == "mousedown on departure, transition: 0--->6"):
        count_mousedown_departure+=1

    if(i == "mousedown on arrival, transition: 0--->5"):
        count_mousedown_arrival+=1

    if(i == "mousedown on airtime, transition: 0--->7"):
        count_mousedown_airtime+=1

    if(i == "mousedown on delay, transition: 0--->8"):
        count_mousedown_delay +=1

    if(i == "mousedown on count, transition: 0--->1"):
        count_mousedown_count +=1
    
    ### mouseup

    if(i == "mouseup on distance, transition: 4--->0"):
        count_mouseup_distance+=1

    if(i == "mouseup on departure, transition: 6--->0"):
        count_mouseup_departure+=1

    if(i == "mouseup on arrival, transition: 5--->0"):
        count_mouseup_arrival+=1

    if(i == "mouseup on airtime, transition: 7--->0"):
        count_mouseup_airtime+=1

    if(i == "mouseup on delay, transition: 8--->0"):
        count_mouseup_delay +=1

    ### click

    if(i == "click on #departure canvas.marks , transition: 0--->0"):
        count_click_departure+=1

    if(i == "click on #distance canvas.marks , transition: 0--->0"):
        count_click_distance+=1

    if(i == "click on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "click on #arrival canvas.marks , transition: 0--->0"):
        count_click_arrival+=1

    if(i == "click on #delay canvas.marks , transition: 0--->0"):
        count_click_delay +=1

    if(i == "click on #count canvas.marks , transition: 0--->0"):
        count_click_count +=1

    if(i == "click on #count canvas.marks , transition: 1--->1"):
        special_click_count +=1

    ### dblclick


    if(i == "dblclick on #departure canvas.marks , transition: 0--->0"):
        count_dblclick_departure+=1

    if(i == "dblclickk on #distance canvas.marks , transition: 0--->0"):
        count_dblclick_distance+=1

    if(i == "dblclick on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "dblclick on #arrival canvas.marks , transition: 0--->0"):
        count_dblclick_arrival+=1

    if(i == "dblclick on #delay canvas.marks , transition: 0--->0"):
        count_dblclick_delay +=1

    if(i == "dblclick on #count canvas.marks , transition: 0--->0"):
        count_dblclick_count +=1

    ### wheel

    if(i == "wheel on #distance canvas.marks , transition: 0--->0"):
        count_wheel_distance +=1
    
    if(i == "wheel on #departure canvas.marks , transition: 0--->0"):
        count_wheel_departure +=1
    
    if(i == "wheel on #airtime canvas.marks , transition: 0--->0"):
        count_wheel_airtime +=1
    
    if(i == "wheel on #arrival canvas.marks , transition: 0--->0"):
        count_wheel_arrival +=1
    
    if(i == "wheel on #delay canvas.marks , transition: 0--->0"):
        count_wheel_delay +=1

    ### special wheel

    if(i == "wheel on departure while brushing, transition: 6--->6"):
        special_wheel_departure +=1
    
    if(i == "wheel on arrival while brushing, transition: 5--->5"):
        special_wheel_arrival +=1
    
    if(i == "wheel on airtime while brushing, transition: 7--->7"):
        special_wheel_airtime +=1
    
    if(i == "wheel on delay while brushing, transition: 8--->8"):
        special_wheel_delay +=1
    
    if(i == "wheel on distance while brushing, transition: 4--->4"):
        special_wheel_distance +=1

    ### mouseout

    if(i == "mouseout on #distance canvas.marks , transition: 0--->0"):
        count_mouseout_distance +=1
    
    if(i == "mouseout on #departure canvas.marks , transition: 0--->0"):
        count_mouseout_departure +=1
    
    if(i == "mouseout on #airtime canvas.marks , transition: 0--->0"):
        count_mouseout_airtime +=1
    
    if(i == "mouseout on #arrival canvas.marks , transition: 0--->0"):
        count_mouseout_arrival +=1
    
    if(i == "mouseout on #delay canvas.marks , transition: 0--->0"):
        count_mouseout_delay +=1
    
    if(i == "mouseout on #count canvas.marks , transition: 0--->0"):
        count_mouseout_count +=1

    if(i == "mouseout on #count canvas.marks , transition: 1--->1"):
        special_mouseout_count +=1

    ### special mouseout

    if(i == "mouseout on departure while brushing, transition: 6--->6"):
        special_mouseout_departure +=1
    
    if(i == "mouseout on arrival while brushing, transition: 5--->5"):
        special_mouseout_arrival +=1
    
    if(i == "mouseout on airtime while brushing, transition: 7--->7"):
        special_mouseout_airtime +=1
    
    if(i == "mouseout on delay while brushing, transition: 8--->8"):
        special_mouseout_delay +=1
    
    if(i == "mouseout on distance while brushing, transition: 4--->4"):
        special_mouseout_distance +=1



exploredNodes.append("----------------------PATH 2------------------")

if(count_mousemove_departure) != 0:exploredNodes.append("number of mousemove on departure, transition: 0--->0 = " + str(count_mousemove_departure))
if(count_mousemove_delay) != 0:exploredNodes.append("number of mousemove on delay, transition: 0--->0 = " + str(count_mousemove_delay))
if(count_mousemove_airtime) != 0:exploredNodes.append("number of mousemove on airtime, transition: 0--->0 = " + str(count_mousemove_airtime))
if(count_mousemove_arrival) != 0:exploredNodes.append("number of mousemove on arrival, transition: 0--->0 = " + str(count_mousemove_arrival))
if(count_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 0--->0 = " + str(count_mousemove_count))
if(count_mousemove_distance) != 0:exploredNodes.append("number of mousemove on distance, transition: 0--->0 = " + str(count_mousemove_distance))

if(count_mousemove_brush_departure) != 0:exploredNodes.append("number of special mousemove on departure, transition: 6--->6 = " + str(count_mousemove_brush_departure))
if(count_mousemove_brush_arrival) != 0:exploredNodes.append("number of special mousemove on arrival, transition: 5--->5 = " + str(count_mousemove_brush_arrival))
if(count_mousemove_brush_airtime) != 0:exploredNodes.append("number of special mousemove on airtime, transition: 7--->7 = " + str(count_mousemove_brush_airtime))
if(count_mousemove_brush_delay) != 0:exploredNodes.append("number of special mousemove on delay, transition: 8--->8 = " + str(count_mousemove_brush_delay))
if(count_mousemove_brush_distance) != 0:exploredNodes.append("number of special mousemove on distance, transition: 4--->4 = " + str(count_mousemove_brush_distance))

if(special_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 1--->1 = " + str(special_mousemove_count))

if(count_mousedown_distance) != 0:exploredNodes.append("number of mousedown on distance, transition: 0--->4 = " + str(count_mousedown_distance))
if(count_mousedown_departure) != 0:exploredNodes.append("number of mousedown on departure, transition: 0--->6 = " + str(count_mousedown_departure))
if(count_mousedown_arrival) != 0:exploredNodes.append("number of mousedown on arrival, transition: 0--->5 = " + str(count_mousedown_arrival))
if(count_mousedown_airtime) != 0:exploredNodes.append("number of mousedown on airtime, transition: 0--->7 = " + str(count_mousedown_airtime))
if(count_mousedown_delay) != 0:exploredNodes.append("number of mousedown on delay, transition: 0--->8 = " + str(count_mousedown_delay))
if(count_mousedown_count) != 0:exploredNodes.append("number of mousedown on count, transition: 0--->1 = " + str(count_mousedown_count))

if(count_mouseup_distance) != 0:exploredNodes.append("number of mouseup on distance, transition: 4--->0 = " + str(count_mouseup_distance))
if(count_mouseup_departure) != 0:exploredNodes.append("number of mouseup on departure, transition: 6--->0 = " + str(count_mouseup_departure))
if(count_mouseup_arrival) != 0:exploredNodes.append("number of mouseup on arrival, transition: 5--->0 = " + str(count_mouseup_arrival))
if(count_mouseup_airtime) != 0:exploredNodes.append("number of mouseup on airtime, transition: 7--->0 = " + str(count_mouseup_airtime))
if(count_mouseup_delay) != 0:exploredNodes.append("number of mouseup on delay, transition: 8--->0 = " + str(count_mouseup_delay))

if(count_click_distance) != 0:exploredNodes.append("number of click on distance, transition: 0--->0 = " + str(count_click_distance))
if(count_click_departure) != 0:exploredNodes.append("number of click on departure, transition: 0--->0 = " + str(count_click_departure))
if(count_click_arrival) != 0:exploredNodes.append("number of click on arrival, transition: 0--->0 = " + str(count_click_arrival))
if(count_click_airtime) != 0:exploredNodes.append("number of click on airtime, transition: 0--->0 = " + str(count_click_airtime))
if(count_click_delay) != 0:exploredNodes.append("number of click on delay, transition: 0--->0 = " + str(count_click_delay))
if(count_click_count) != 0:exploredNodes.append("number of click on count, transition: 0--->0 = " + str(count_click_count))
if(special_click_count) != 0:exploredNodes.append("number of special click on count, transition: 1--->1 = " + str(special_click_count))

if(count_dblclick_distance) != 0:exploredNodes.append("number of dblclick on distance, transition: 0--->0 = " + str(count_dblclick_distance))
if(count_dblclick_departure) != 0:exploredNodes.append("number of dblclick on departure, transition: 0--->0 = " + str(count_dblclick_departure))
if(count_dblclick_arrival) != 0:exploredNodes.append("number of dblclick on arrival, transition: 0--->0 = " + str(count_dblclick_arrival))
if(count_dblclick_airtime) != 0:exploredNodes.append("number of dblclick on airtime, transition: 0--->0 = " + str(count_dblclick_airtime))
if(count_dblclick_delay) != 0:exploredNodes.append("number of dblclick on delay, transition: 0--->0 = " + str(count_dblclick_delay))
if(count_dblclick_count) != 0:exploredNodes.append("number of dblclick on count, transition: 0--->0 = " + str(count_dblclick_count))

if(count_wheel_distance) != 0:exploredNodes.append("number of wheel on distance, transition: 0--->0 = " + str(count_wheel_distance))
if(count_wheel_departure) != 0:exploredNodes.append("number of wheel on departure, transition: 0--->0 = " + str(count_wheel_departure))
if(count_wheel_arrival) != 0:exploredNodes.append("number of wheel on arrival, transition: 0---> = " + str(count_wheel_arrival))
if(count_wheel_airtime) != 0:exploredNodes.append("number of wheel on airtime, transition: 0--->0 = " + str(count_wheel_airtime))
if(count_wheel_delay) != 0:exploredNodes.append("number of wheel on delay, transition: 0--->0 = " + str(count_wheel_delay))

if(special_wheel_distance) != 0:exploredNodes.append("number of special wheel on distance, transition: 4--->4 = " + str(special_wheel_distance))
if(special_wheel_departure) != 0:exploredNodes.append("number of special wheel on departure, transition: 6--->6 = " + str(special_wheel_departure))
if(special_wheel_arrival) != 0:exploredNodes.append("number of special wheel on arrival, transition: 5--->5 = " + str(special_wheel_arrival))
if(special_wheel_airtime) != 0:exploredNodes.append("number of special wheel on airtime, transition: 7--->7 = " + str(special_wheel_airtime))
if(special_wheel_delay) != 0:exploredNodes.append("number of special wheel on delay, transition: 8--->8 = " + str(special_wheel_delay))

if(count_mouseout_distance) != 0:exploredNodes.append("number of mouseout on distance, transition: 0--->0 = " + str(count_mouseout_distance))
if(count_mouseout_departure) != 0:exploredNodes.append("number of mouseout on departure, transition: 0--->0 = " + str(count_mouseout_departure))
if(count_mouseout_arrival) != 0:exploredNodes.append("number of mouseout on arrival, transition: 0--->0 = " + str(count_mouseout_arrival))
if(count_mouseout_airtime) != 0:exploredNodes.append("number of mouseout on airtime, transition: 0--->0 = " + str(count_mouseout_airtime))
if(count_mouseout_delay) != 0:exploredNodes.append("number of mouseout on delay, transition: 0--->0 = " + str(count_mouseout_delay))

if(special_mouseout_distance) != 0:exploredNodes.append("number of special mouseout on distance, transition: 4--->4 = " + str(special_mouseout_distance))
if(special_mouseout_departure) != 0:exploredNodes.append("number of special mouseout on departure, transition: 6--->6 = " + str(special_mouseout_departure))
if(special_mouseout_arrival) != 0:exploredNodes.append("number of special mouseout on arrival, transition: 5--->5 = " + str(special_mouseout_arrival))
if(special_mouseout_airtime) != 0:exploredNodes.append("number of special mouseout on airtime, transition: 7--->7 = " + str(special_mouseout_airtime))
if(special_mouseout_delay) != 0:exploredNodes.append("number of special mouseout on delay, transition: 8--->8 = " + str(special_mouseout_delay))
if(special_mouseout_count) != 0:exploredNodes.append("number of special mouseout on count, transition: 1--->1 = " + str(special_mouseout_count))


#save on excel file

sheet_total["A3"].value += count_mousemove_departure
sheet_total["B3"].value += count_mousemove_distance
sheet_total["C3"].value += count_mousemove_delay
sheet_total["D3"].value += count_mousemove_airtime
sheet_total["E3"].value += count_mousemove_arrival
sheet_total["F3"].value += count_mousemove_count
sheet_total["G3"].value += special_mousemove_count

sheet_total["A6"].value += count_mousedown_departure
sheet_total["B6"].value += count_mousedown_distance
sheet_total["C6"].value += count_mousedown_delay
sheet_total["D6"].value += count_mousedown_airtime
sheet_total["E6"].value += count_mousedown_arrival
sheet_total["F6"].value += count_mousedown_count

sheet_total["A10"].value += count_mousemove_brush_departure
sheet_total["B10"].value += count_mousemove_brush_distance
sheet_total["C10"].value += count_mousemove_brush_delay
sheet_total["D10"].value += count_mousemove_brush_airtime
sheet_total["E10"].value += count_mousemove_brush_arrival

sheet_total["A13"].value += count_mouseup_departure
sheet_total["B13"].value += count_mouseup_distance
sheet_total["C13"].value += count_mouseup_delay
sheet_total["D13"].value += count_mouseup_airtime
sheet_total["E13"].value += count_mouseup_arrival

sheet_total["A17"].value += count_click_departure
sheet_total["B17"].value += count_click_distance
sheet_total["C17"].value += count_click_delay
sheet_total["D17"].value += count_click_airtime
sheet_total["E17"].value += count_click_arrival
sheet_total["F17"].value += count_click_count
sheet_total["G17"].value += special_click_count

sheet_total["A21"].value += count_mouseout_departure
sheet_total["B21"].value += count_mouseout_distance
sheet_total["C21"].value += count_mouseout_delay
sheet_total["D21"].value += count_mouseout_airtime
sheet_total["E21"].value += count_mouseout_arrival
sheet_total["F21"].value += count_mouseout_count

sheet_total["A25"].value += special_mouseout_departure
sheet_total["B25"].value += special_mouseout_distance
sheet_total["C25"].value += special_mouseout_delay
sheet_total["D25"].value += special_mouseout_airtime
sheet_total["E25"].value += special_mouseout_arrival
sheet_total["F25"].value += special_mouseout_count

sheet_total["A29"].value += count_wheel_departure
sheet_total["B29"].value += count_wheel_distance
sheet_total["C29"].value += count_wheel_delay
sheet_total["D29"].value += count_wheel_airtime
sheet_total["E29"].value += count_wheel_arrival


sheet_total["A33"].value += special_wheel_departure
sheet_total["B33"].value += special_wheel_distance
sheet_total["C33"].value += special_wheel_delay
sheet_total["D33"].value += special_wheel_airtime
sheet_total["E33"].value += special_wheel_arrival

sheet_total["A37"].value += count_dblclick_departure
sheet_total["B37"].value += count_dblclick_distance
sheet_total["C37"].value += count_dblclick_delay
sheet_total["D37"].value += count_dblclick_airtime
sheet_total["E37"].value += count_dblclick_arrival
sheet_total["F37"].value += count_dblclick_count

wb_total.save(filename="total_count.xlsx")

sheet_2["A3"].value += count_mousemove_departure
sheet_2["B3"].value += count_mousemove_distance
sheet_2["C3"].value += count_mousemove_delay
sheet_2["D3"].value += count_mousemove_airtime
sheet_2["E3"].value += count_mousemove_arrival
sheet_2["F3"].value += count_mousemove_count
sheet_2["G3"].value += special_mousemove_count

sheet_2["A6"].value += count_mousedown_departure
sheet_2["B6"].value += count_mousedown_distance
sheet_2["C6"].value += count_mousedown_delay
sheet_2["D6"].value += count_mousedown_airtime
sheet_2["E6"].value += count_mousedown_arrival
sheet_2["F6"].value += count_mousedown_count

sheet_2["A10"].value += count_mousemove_brush_departure
sheet_2["B10"].value += count_mousemove_brush_distance
sheet_2["C10"].value += count_mousemove_brush_delay
sheet_2["D10"].value += count_mousemove_brush_airtime
sheet_2["E10"].value += count_mousemove_brush_arrival

sheet_2["A13"].value += count_mouseup_departure
sheet_2["B13"].value += count_mouseup_distance
sheet_2["C13"].value += count_mouseup_delay
sheet_2["D13"].value += count_mouseup_airtime
sheet_2["E13"].value += count_mouseup_arrival

sheet_2["A17"].value += count_click_departure
sheet_2["B17"].value += count_click_distance
sheet_2["C17"].value += count_click_delay
sheet_2["D17"].value += count_click_airtime
sheet_2["E17"].value += count_click_arrival
sheet_2["F17"].value += count_click_count
sheet_2["G17"].value += special_click_count

sheet_2["A21"].value += count_mouseout_departure
sheet_2["B21"].value += count_mouseout_distance
sheet_2["C21"].value += count_mouseout_delay
sheet_2["D21"].value += count_mouseout_airtime
sheet_2["E21"].value += count_mouseout_arrival
sheet_2["F21"].value += count_mouseout_count

sheet_2["A25"].value += special_mouseout_departure
sheet_2["B25"].value += special_mouseout_distance
sheet_2["C25"].value += special_mouseout_delay
sheet_2["D25"].value += special_mouseout_airtime
sheet_2["E25"].value += special_mouseout_arrival
sheet_2["F25"].value += special_mouseout_count

sheet_2["A29"].value += count_wheel_departure
sheet_2["B29"].value += count_wheel_distance
sheet_2["C29"].value += count_wheel_delay
sheet_2["D29"].value += count_wheel_airtime
sheet_2["E29"].value += count_wheel_arrival


sheet_2["A33"].value += special_wheel_departure
sheet_2["B33"].value += special_wheel_distance
sheet_2["C33"].value += special_wheel_delay
sheet_2["D33"].value += special_wheel_airtime
sheet_2["E33"].value += special_wheel_arrival

sheet_2["A37"].value += count_dblclick_departure
sheet_2["B37"].value += count_dblclick_distance
sheet_2["C37"].value += count_dblclick_delay
sheet_2["D37"].value += count_dblclick_airtime
sheet_2["E37"].value += count_dblclick_arrival
sheet_2["F37"].value += count_dblclick_count

wb_2.save(filename="count_path_2.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

for i in data["2"]:
    ### mousemove
    
    if(i == "mousemove on #distance canvas.marks , transition: 0--->0"):
        count_mousemove_distance = count_mousemove_distance + 1
    
    if(i == "mousemove on #arrival canvas.marks , transition: 0--->0"):
        count_mousemove_arrival = count_mousemove_arrival + 1
    
    if(i == "mousemove on #count canvas.marks , transition: 0--->0"):
        count_mousemove_count = count_mousemove_count + 1
    
    if(i == "mousemove on #departure canvas.marks , transition: 0--->0"):
        count_mousemove_departure = count_mousemove_departure + 1
    
    if(i == "mousemove on #airtime canvas.marks , transition: 0--->0"):
        count_mousemove_airtime = count_mousemove_airtime + 1
    
    if(i == "mousemove on #delay canvas.marks , transition: 0--->0"):
        count_mousemove_delay = count_mousemove_delay + 1
        
        ### mousemove brushing

    if(i == "mousemove on departure, transition: 6--->6"):
        count_mousemove_brush_departure +=1
    
    if(i == "mousemove on arrival, transition: 5--->5"):
        count_mousemove_brush_arrival +=1
    
    if(i == "mousemove on airtime, transition: 7--->7"):
        count_mousemove_brush_airtime +=1
    
    if(i == "mousemove on delay, transition: 8--->8"):
        count_mousemove_brush_delay +=1
    
    if(i == "mousemove on distance, transition: 4--->4"):
        count_mousemove_brush_distance +=1

    ### mousemove count

    if(i == "mousemove on count, transition: 1--->1"):
        special_mousemove_count+=1

    ### mousedown

    if(i == "mousedown on distance, transition: 0--->4"):
        count_mousedown_distance+=1

    if(i == "mousedown on departure, transition: 0--->6"):
        count_mousedown_departure+=1

    if(i == "mousedown on arrival, transition: 0--->5"):
        count_mousedown_arrival+=1

    if(i == "mousedown on airtime, transition: 0--->7"):
        count_mousedown_airtime+=1

    if(i == "mousedown on delay, transition: 0--->8"):
        count_mousedown_delay +=1

    if(i == "mousedown on count, transition: 0--->1"):
        count_mousedown_count +=1
    
    ### mouseup

    if(i == "mouseup on distance, transition: 4--->0"):
        count_mouseup_distance+=1

    if(i == "mouseup on departure, transition: 6--->0"):
        count_mouseup_departure+=1

    if(i == "mouseup on arrival, transition: 5--->0"):
        count_mouseup_arrival+=1

    if(i == "mouseup on airtime, transition: 7--->0"):
        count_mouseup_airtime+=1

    if(i == "mouseup on delay, transition: 8--->0"):
        count_mouseup_delay +=1

    ### click

    if(i == "click on #departure canvas.marks , transition: 0--->0"):
        count_click_departure+=1

    if(i == "click on #distance canvas.marks , transition: 0--->0"):
        count_click_distance+=1

    if(i == "click on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "click on #arrival canvas.marks , transition: 0--->0"):
        count_click_arrival+=1

    if(i == "click on #delay canvas.marks , transition: 0--->0"):
        count_click_delay +=1

    if(i == "click on #count canvas.marks , transition: 0--->0"):
        count_click_count +=1

    if(i == "click on #count canvas.marks , transition: 1--->1"):
        special_click_count +=1

    ### dblclick


    if(i == "dblclick on #departure canvas.marks , transition: 0--->0"):
        count_dblclick_departure+=1

    if(i == "dblclickk on #distance canvas.marks , transition: 0--->0"):
        count_dblclick_distance+=1

    if(i == "dblclick on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "dblclick on #arrival canvas.marks , transition: 0--->0"):
        count_dblclick_arrival+=1

    if(i == "dblclick on #delay canvas.marks , transition: 0--->0"):
        count_dblclick_delay +=1

    if(i == "dblclick on #count canvas.marks , transition: 0--->0"):
        count_dblclick_count +=1

    ### wheel

    if(i == "wheel on #distance canvas.marks , transition: 0--->0"):
        count_wheel_distance +=1
    
    if(i == "wheel on #departure canvas.marks , transition: 0--->0"):
        count_wheel_departure +=1
    
    if(i == "wheel on #airtime canvas.marks , transition: 0--->0"):
        count_wheel_airtime +=1
    
    if(i == "wheel on #arrival canvas.marks , transition: 0--->0"):
        count_wheel_arrival +=1
    
    if(i == "wheel on #delay canvas.marks , transition: 0--->0"):
        count_wheel_delay +=1

    ### special wheel

    if(i == "wheel on departure while brushing, transition: 6--->6"):
        special_wheel_departure +=1
    
    if(i == "wheel on arrival while brushing, transition: 5--->5"):
        special_wheel_arrival +=1
    
    if(i == "wheel on airtime while brushing, transition: 7--->7"):
        special_wheel_airtime +=1
    
    if(i == "wheel on delay while brushing, transition: 8--->8"):
        special_wheel_delay +=1
    
    if(i == "wheel on distance while brushing, transition: 4--->4"):
        special_wheel_distance +=1

    ### mouseout

    if(i == "mouseout on #distance canvas.marks , transition: 0--->0"):
        count_mouseout_distance +=1
    
    if(i == "mouseout on #departure canvas.marks , transition: 0--->0"):
        count_mouseout_departure +=1
    
    if(i == "mouseout on #airtime canvas.marks , transition: 0--->0"):
        count_mouseout_airtime +=1
    
    if(i == "mouseout on #arrival canvas.marks , transition: 0--->0"):
        count_mouseout_arrival +=1
    
    if(i == "mouseout on #delay canvas.marks , transition: 0--->0"):
        count_mouseout_delay +=1
    
    if(i == "mouseout on #count canvas.marks , transition: 0--->0"):
        count_mouseout_count +=1

    if(i == "mouseout on #count canvas.marks , transition: 1--->1"):
        special_mouseout_count +=1

    ### special mouseout

    if(i == "mouseout on departure while brushing, transition: 6--->6"):
        special_mouseout_departure +=1
    
    if(i == "mouseout on arrival while brushing, transition: 5--->5"):
        special_mouseout_arrival +=1
    
    if(i == "mouseout on airtime while brushing, transition: 7--->7"):
        special_mouseout_airtime +=1
    
    if(i == "mouseout on delay while brushing, transition: 8--->8"):
        special_mouseout_delay +=1
    
    if(i == "mouseout on distance while brushing, transition: 4--->4"):
        special_mouseout_distance +=1



exploredNodes.append("----------------------PATH 3------------------")

if(count_mousemove_departure) != 0:exploredNodes.append("number of mousemove on departure, transition: 0--->0 = " + str(count_mousemove_departure))
if(count_mousemove_delay) != 0:exploredNodes.append("number of mousemove on delay, transition: 0--->0 = " + str(count_mousemove_delay))
if(count_mousemove_airtime) != 0:exploredNodes.append("number of mousemove on airtime, transition: 0--->0 = " + str(count_mousemove_airtime))
if(count_mousemove_arrival) != 0:exploredNodes.append("number of mousemove on arrival, transition: 0--->0 = " + str(count_mousemove_arrival))
if(count_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 0--->0 = " + str(count_mousemove_count))
if(count_mousemove_distance) != 0:exploredNodes.append("number of mousemove on distance, transition: 0--->0 = " + str(count_mousemove_distance))

if(count_mousemove_brush_departure) != 0:exploredNodes.append("number of special mousemove on departure, transition: 6--->6 = " + str(count_mousemove_brush_departure))
if(count_mousemove_brush_arrival) != 0:exploredNodes.append("number of special mousemove on arrival, transition: 5--->5 = " + str(count_mousemove_brush_arrival))
if(count_mousemove_brush_airtime) != 0:exploredNodes.append("number of special mousemove on airtime, transition: 7--->7 = " + str(count_mousemove_brush_airtime))
if(count_mousemove_brush_delay) != 0:exploredNodes.append("number of special mousemove on delay, transition: 8--->8 = " + str(count_mousemove_brush_delay))
if(count_mousemove_brush_distance) != 0:exploredNodes.append("number of special mousemove on distance, transition: 4--->4 = " + str(count_mousemove_brush_distance))

if(special_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 1--->1 = " + str(special_mousemove_count))

if(count_mousedown_distance) != 0:exploredNodes.append("number of mousedown on distance, transition: 0--->4 = " + str(count_mousedown_distance))
if(count_mousedown_departure) != 0:exploredNodes.append("number of mousedown on departure, transition: 0--->6 = " + str(count_mousedown_departure))
if(count_mousedown_arrival) != 0:exploredNodes.append("number of mousedown on arrival, transition: 0--->5 = " + str(count_mousedown_arrival))
if(count_mousedown_airtime) != 0:exploredNodes.append("number of mousedown on airtime, transition: 0--->7 = " + str(count_mousedown_airtime))
if(count_mousedown_delay) != 0:exploredNodes.append("number of mousedown on delay, transition: 0--->8 = " + str(count_mousedown_delay))
if(count_mousedown_count) != 0:exploredNodes.append("number of mousedown on count, transition: 0--->1 = " + str(count_mousedown_count))

if(count_mouseup_distance) != 0:exploredNodes.append("number of mouseup on distance, transition: 4--->0 = " + str(count_mouseup_distance))
if(count_mouseup_departure) != 0:exploredNodes.append("number of mouseup on departure, transition: 6--->0 = " + str(count_mouseup_departure))
if(count_mouseup_arrival) != 0:exploredNodes.append("number of mouseup on arrival, transition: 5--->0 = " + str(count_mouseup_arrival))
if(count_mouseup_airtime) != 0:exploredNodes.append("number of mouseup on airtime, transition: 7--->0 = " + str(count_mouseup_airtime))
if(count_mouseup_delay) != 0:exploredNodes.append("number of mouseup on delay, transition: 8--->0 = " + str(count_mouseup_delay))

if(count_click_distance) != 0:exploredNodes.append("number of click on distance, transition: 0--->0 = " + str(count_click_distance))
if(count_click_departure) != 0:exploredNodes.append("number of click on departure, transition: 0--->0 = " + str(count_click_departure))
if(count_click_arrival) != 0:exploredNodes.append("number of click on arrival, transition: 0--->0 = " + str(count_click_arrival))
if(count_click_airtime) != 0:exploredNodes.append("number of click on airtime, transition: 0--->0 = " + str(count_click_airtime))
if(count_click_delay) != 0:exploredNodes.append("number of click on delay, transition: 0--->0 = " + str(count_click_delay))
if(count_click_count) != 0:exploredNodes.append("number of click on count, transition: 0--->0 = " + str(count_click_count))
if(special_click_count) != 0:exploredNodes.append("number of special click on count, transition: 1--->1 = " + str(special_click_count))

if(count_dblclick_distance) != 0:exploredNodes.append("number of dblclick on distance, transition: 0--->0 = " + str(count_dblclick_distance))
if(count_dblclick_departure) != 0:exploredNodes.append("number of dblclick on departure, transition: 0--->0 = " + str(count_dblclick_departure))
if(count_dblclick_arrival) != 0:exploredNodes.append("number of dblclick on arrival, transition: 0--->0 = " + str(count_dblclick_arrival))
if(count_dblclick_airtime) != 0:exploredNodes.append("number of dblclick on airtime, transition: 0--->0 = " + str(count_dblclick_airtime))
if(count_dblclick_delay) != 0:exploredNodes.append("number of dblclick on delay, transition: 0--->0 = " + str(count_dblclick_delay))
if(count_dblclick_count) != 0:exploredNodes.append("number of dblclick on count, transition: 0--->0 = " + str(count_dblclick_count))

if(count_wheel_distance) != 0:exploredNodes.append("number of wheel on distance, transition: 0--->0 = " + str(count_wheel_distance))
if(count_wheel_departure) != 0:exploredNodes.append("number of wheel on departure, transition: 0--->0 = " + str(count_wheel_departure))
if(count_wheel_arrival) != 0:exploredNodes.append("number of wheel on arrival, transition: 0---> = " + str(count_wheel_arrival))
if(count_wheel_airtime) != 0:exploredNodes.append("number of wheel on airtime, transition: 0--->0 = " + str(count_wheel_airtime))
if(count_wheel_delay) != 0:exploredNodes.append("number of wheel on delay, transition: 0--->0 = " + str(count_wheel_delay))

if(special_wheel_distance) != 0:exploredNodes.append("number of special wheel on distance, transition: 4--->4 = " + str(special_wheel_distance))
if(special_wheel_departure) != 0:exploredNodes.append("number of special wheel on departure, transition: 6--->6 = " + str(special_wheel_departure))
if(special_wheel_arrival) != 0:exploredNodes.append("number of special wheel on arrival, transition: 5--->5 = " + str(special_wheel_arrival))
if(special_wheel_airtime) != 0:exploredNodes.append("number of special wheel on airtime, transition: 7--->7 = " + str(special_wheel_airtime))
if(special_wheel_delay) != 0:exploredNodes.append("number of special wheel on delay, transition: 8--->8 = " + str(special_wheel_delay))

if(count_mouseout_distance) != 0:exploredNodes.append("number of mouseout on distance, transition: 0--->0 = " + str(count_mouseout_distance))
if(count_mouseout_departure) != 0:exploredNodes.append("number of mouseout on departure, transition: 0--->0 = " + str(count_mouseout_departure))
if(count_mouseout_arrival) != 0:exploredNodes.append("number of mouseout on arrival, transition: 0--->0 = " + str(count_mouseout_arrival))
if(count_mouseout_airtime) != 0:exploredNodes.append("number of mouseout on airtime, transition: 0--->0 = " + str(count_mouseout_airtime))
if(count_mouseout_delay) != 0:exploredNodes.append("number of mouseout on delay, transition: 0--->0 = " + str(count_mouseout_delay))

if(special_mouseout_distance) != 0:exploredNodes.append("number of special mouseout on distance, transition: 4--->4 = " + str(special_mouseout_distance))
if(special_mouseout_departure) != 0:exploredNodes.append("number of special mouseout on departure, transition: 6--->6 = " + str(special_mouseout_departure))
if(special_mouseout_arrival) != 0:exploredNodes.append("number of special mouseout on arrival, transition: 5--->5 = " + str(special_mouseout_arrival))
if(special_mouseout_airtime) != 0:exploredNodes.append("number of special mouseout on airtime, transition: 7--->7 = " + str(special_mouseout_airtime))
if(special_mouseout_delay) != 0:exploredNodes.append("number of special mouseout on delay, transition: 8--->8 = " + str(special_mouseout_delay))
if(special_mouseout_count) != 0:exploredNodes.append("number of special mouseout on count, transition: 1--->1 = " + str(special_mouseout_count))

#save on excel file

sheet_total["A3"].value += count_mousemove_departure
sheet_total["B3"].value += count_mousemove_distance
sheet_total["C3"].value += count_mousemove_delay
sheet_total["D3"].value += count_mousemove_airtime
sheet_total["E3"].value += count_mousemove_arrival
sheet_total["F3"].value += count_mousemove_count
sheet_total["G3"].value += special_mousemove_count

sheet_total["A6"].value += count_mousedown_departure
sheet_total["B6"].value += count_mousedown_distance
sheet_total["C6"].value += count_mousedown_delay
sheet_total["D6"].value += count_mousedown_airtime
sheet_total["E6"].value += count_mousedown_arrival
sheet_total["F6"].value += count_mousedown_count

sheet_total["A10"].value += count_mousemove_brush_departure
sheet_total["B10"].value += count_mousemove_brush_distance
sheet_total["C10"].value += count_mousemove_brush_delay
sheet_total["D10"].value += count_mousemove_brush_airtime
sheet_total["E10"].value += count_mousemove_brush_arrival

sheet_total["A13"].value += count_mouseup_departure
sheet_total["B13"].value += count_mouseup_distance
sheet_total["C13"].value += count_mouseup_delay
sheet_total["D13"].value += count_mouseup_airtime
sheet_total["E13"].value += count_mouseup_arrival

sheet_total["A17"].value += count_click_departure
sheet_total["B17"].value += count_click_distance
sheet_total["C17"].value += count_click_delay
sheet_total["D17"].value += count_click_airtime
sheet_total["E17"].value += count_click_arrival
sheet_total["F17"].value += count_click_count
sheet_total["G17"].value += special_click_count

sheet_total["A21"].value += count_mouseout_departure
sheet_total["B21"].value += count_mouseout_distance
sheet_total["C21"].value += count_mouseout_delay
sheet_total["D21"].value += count_mouseout_airtime
sheet_total["E21"].value += count_mouseout_arrival
sheet_total["F21"].value += count_mouseout_count

sheet_total["A25"].value += special_mouseout_departure
sheet_total["B25"].value += special_mouseout_distance
sheet_total["C25"].value += special_mouseout_delay
sheet_total["D25"].value += special_mouseout_airtime
sheet_total["E25"].value += special_mouseout_arrival
sheet_total["F25"].value += special_mouseout_count

sheet_total["A29"].value += count_wheel_departure
sheet_total["B29"].value += count_wheel_distance
sheet_total["C29"].value += count_wheel_delay
sheet_total["D29"].value += count_wheel_airtime
sheet_total["E29"].value += count_wheel_arrival


sheet_total["A33"].value += special_wheel_departure
sheet_total["B33"].value += special_wheel_distance
sheet_total["C33"].value += special_wheel_delay
sheet_total["D33"].value += special_wheel_airtime
sheet_total["E33"].value += special_wheel_arrival

sheet_total["A37"].value += count_dblclick_departure
sheet_total["B37"].value += count_dblclick_distance
sheet_total["C37"].value += count_dblclick_delay
sheet_total["D37"].value += count_dblclick_airtime
sheet_total["E37"].value += count_dblclick_arrival
sheet_total["F37"].value += count_dblclick_count

wb_total.save(filename="total_count.xlsx")

sheet_3["A3"].value += count_mousemove_departure
sheet_3["B3"].value += count_mousemove_distance
sheet_3["C3"].value += count_mousemove_delay
sheet_3["D3"].value += count_mousemove_airtime
sheet_3["E3"].value += count_mousemove_arrival
sheet_3["F3"].value += count_mousemove_count
sheet_3["G3"].value += special_mousemove_count

sheet_3["A6"].value += count_mousedown_departure
sheet_3["B6"].value += count_mousedown_distance
sheet_3["C6"].value += count_mousedown_delay
sheet_3["D6"].value += count_mousedown_airtime
sheet_3["E6"].value += count_mousedown_arrival
sheet_3["F6"].value += count_mousedown_count

sheet_3["A10"].value += count_mousemove_brush_departure
sheet_3["B10"].value += count_mousemove_brush_distance
sheet_3["C10"].value += count_mousemove_brush_delay
sheet_3["D10"].value += count_mousemove_brush_airtime
sheet_3["E10"].value += count_mousemove_brush_arrival

sheet_3["A13"].value += count_mouseup_departure
sheet_3["B13"].value += count_mouseup_distance
sheet_3["C13"].value += count_mouseup_delay
sheet_3["D13"].value += count_mouseup_airtime
sheet_3["E13"].value += count_mouseup_arrival

sheet_3["A17"].value += count_click_departure
sheet_3["B17"].value += count_click_distance
sheet_3["C17"].value += count_click_delay
sheet_3["D17"].value += count_click_airtime
sheet_3["E17"].value += count_click_arrival
sheet_3["F17"].value += count_click_count
sheet_3["G17"].value += special_click_count

sheet_3["A21"].value += count_mouseout_departure
sheet_3["B21"].value += count_mouseout_distance
sheet_3["C21"].value += count_mouseout_delay
sheet_3["D21"].value += count_mouseout_airtime
sheet_3["E21"].value += count_mouseout_arrival
sheet_3["F21"].value += count_mouseout_count

sheet_3["A25"].value += special_mouseout_departure
sheet_3["B25"].value += special_mouseout_distance
sheet_3["C25"].value += special_mouseout_delay
sheet_3["D25"].value += special_mouseout_airtime
sheet_3["E25"].value += special_mouseout_arrival
sheet_3["F25"].value += special_mouseout_count

sheet_3["A29"].value += count_wheel_departure
sheet_3["B29"].value += count_wheel_distance
sheet_3["C29"].value += count_wheel_delay
sheet_3["D29"].value += count_wheel_airtime
sheet_3["E29"].value += count_wheel_arrival


sheet_3["A33"].value += special_wheel_departure
sheet_3["B33"].value += special_wheel_distance
sheet_3["C33"].value += special_wheel_delay
sheet_3["D33"].value += special_wheel_airtime
sheet_3["E33"].value += special_wheel_arrival

sheet_3["A37"].value += count_dblclick_departure
sheet_3["B37"].value += count_dblclick_distance
sheet_3["C37"].value += count_dblclick_delay
sheet_3["D37"].value += count_dblclick_airtime
sheet_3["E37"].value += count_dblclick_arrival
sheet_3["F37"].value += count_dblclick_count

wb_3.save(filename="count_path_3.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

for i in data["3"]:

    ### mousemove
    
    if(i == "mousemove on #distance canvas.marks , transition: 0--->0"):
        count_mousemove_distance = count_mousemove_distance + 1
    
    if(i == "mousemove on #arrival canvas.marks , transition: 0--->0"):
        count_mousemove_arrival = count_mousemove_arrival + 1
    
    if(i == "mousemove on #count canvas.marks , transition: 0--->0"):
        count_mousemove_count = count_mousemove_count + 1
    
    if(i == "mousemove on #departure canvas.marks , transition: 0--->0"):
        count_mousemove_departure = count_mousemove_departure + 1
    
    if(i == "mousemove on #airtime canvas.marks , transition: 0--->0"):
        count_mousemove_airtime = count_mousemove_airtime + 1
    
    if(i == "mousemove on #delay canvas.marks , transition: 0--->0"):
        count_mousemove_delay = count_mousemove_delay + 1
        
        ### mousemove brushing

    if(i == "mousemove on departure, transition: 6--->6"):
        count_mousemove_brush_departure +=1
    
    if(i == "mousemove on arrival, transition: 5--->5"):
        count_mousemove_brush_arrival +=1
    
    if(i == "mousemove on airtime, transition: 7--->7"):
        count_mousemove_brush_airtime +=1
    
    if(i == "mousemove on delay, transition: 8--->8"):
        count_mousemove_brush_delay +=1
    
    if(i == "mousemove on distance, transition: 4--->4"):
        count_mousemove_brush_distance +=1

    ### mousemove count

    if(i == "mousemove on count, transition: 1--->1"):
        special_mousemove_count+=1

    ### mousedown

    if(i == "mousedown on distance, transition: 0--->4"):
        count_mousedown_distance+=1

    if(i == "mousedown on departure, transition: 0--->6"):
        count_mousedown_departure+=1

    if(i == "mousedown on arrival, transition: 0--->5"):
        count_mousedown_arrival+=1

    if(i == "mousedown on airtime, transition: 0--->7"):
        count_mousedown_airtime+=1

    if(i == "mousedown on delay, transition: 0--->8"):
        count_mousedown_delay +=1

    if(i == "mousedown on count, transition: 0--->1"):
        count_mousedown_count +=1
    
    ### mouseup

    if(i == "mouseup on distance, transition: 4--->0"):
        count_mouseup_distance+=1

    if(i == "mouseup on departure, transition: 6--->0"):
        count_mouseup_departure+=1

    if(i == "mouseup on arrival, transition: 5--->0"):
        count_mouseup_arrival+=1

    if(i == "mouseup on airtime, transition: 7--->0"):
        count_mouseup_airtime+=1

    if(i == "mouseup on delay, transition: 8--->0"):
        count_mouseup_delay +=1

    ### click

    if(i == "click on #departure canvas.marks , transition: 0--->0"):
        count_click_departure+=1

    if(i == "click on #distance canvas.marks , transition: 0--->0"):
        count_click_distance+=1

    if(i == "click on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "click on #arrival canvas.marks , transition: 0--->0"):
        count_click_arrival+=1

    if(i == "click on #delay canvas.marks , transition: 0--->0"):
        count_click_delay +=1

    if(i == "click on #count canvas.marks , transition: 0--->0"):
        count_click_count +=1

    if(i == "click on #count canvas.marks , transition: 1--->1"):
        special_click_count +=1

    ### dblclick


    if(i == "dblclick on #departure canvas.marks , transition: 0--->0"):
        count_dblclick_departure+=1

    if(i == "dblclickk on #distance canvas.marks , transition: 0--->0"):
        count_dblclick_distance+=1

    if(i == "dblclick on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "dblclick on #arrival canvas.marks , transition: 0--->0"):
        count_dblclick_arrival+=1

    if(i == "dblclick on #delay canvas.marks , transition: 0--->0"):
        count_dblclick_delay +=1

    if(i == "dblclick on #count canvas.marks , transition: 0--->0"):
        count_dblclick_count +=1

    ### wheel

    if(i == "wheel on #distance canvas.marks , transition: 0--->0"):
        count_wheel_distance +=1
    
    if(i == "wheel on #departure canvas.marks , transition: 0--->0"):
        count_wheel_departure +=1
    
    if(i == "wheel on #airtime canvas.marks , transition: 0--->0"):
        count_wheel_airtime +=1
    
    if(i == "wheel on #arrival canvas.marks , transition: 0--->0"):
        count_wheel_arrival +=1
    
    if(i == "wheel on #delay canvas.marks , transition: 0--->0"):
        count_wheel_delay +=1

    ### special wheel

    if(i == "wheel on departure while brushing, transition: 6--->6"):
        special_wheel_departure +=1
    
    if(i == "wheel on arrival while brushing, transition: 5--->5"):
        special_wheel_arrival +=1
    
    if(i == "wheel on airtime while brushing, transition: 7--->7"):
        special_wheel_airtime +=1
    
    if(i == "wheel on delay while brushing, transition: 8--->8"):
        special_wheel_delay +=1
    
    if(i == "wheel on distance while brushing, transition: 4--->4"):
        special_wheel_distance +=1

    ### mouseout

    if(i == "mouseout on #distance canvas.marks , transition: 0--->0"):
        count_mouseout_distance +=1
    
    if(i == "mouseout on #departure canvas.marks , transition: 0--->0"):
        count_mouseout_departure +=1
    
    if(i == "mouseout on #airtime canvas.marks , transition: 0--->0"):
        count_mouseout_airtime +=1
    
    if(i == "mouseout on #arrival canvas.marks , transition: 0--->0"):
        count_mouseout_arrival +=1
    
    if(i == "mouseout on #delay canvas.marks , transition: 0--->0"):
        count_mouseout_delay +=1
    
    if(i == "mouseout on #count canvas.marks , transition: 0--->0"):
        count_mouseout_count +=1

    if(i == "mouseout on #count canvas.marks , transition: 1--->1"):
        special_mouseout_count +=1

    ### special mouseout

    if(i == "mouseout on departure while brushing, transition: 6--->6"):
        special_mouseout_departure +=1
    
    if(i == "mouseout on arrival while brushing, transition: 5--->5"):
        special_mouseout_arrival +=1
    
    if(i == "mouseout on airtime while brushing, transition: 7--->7"):
        special_mouseout_airtime +=1
    
    if(i == "mouseout on delay while brushing, transition: 8--->8"):
        special_mouseout_delay +=1
    
    if(i == "mouseout on distance while brushing, transition: 4--->4"):
        special_mouseout_distance +=1



exploredNodes.append("----------------------PATH 4------------------")

if(count_mousemove_departure) != 0:exploredNodes.append("number of mousemove on departure, transition: 0--->0 = " + str(count_mousemove_departure))
if(count_mousemove_delay) != 0:exploredNodes.append("number of mousemove on delay, transition: 0--->0 = " + str(count_mousemove_delay))
if(count_mousemove_airtime) != 0:exploredNodes.append("number of mousemove on airtime, transition: 0--->0 = " + str(count_mousemove_airtime))
if(count_mousemove_arrival) != 0:exploredNodes.append("number of mousemove on arrival, transition: 0--->0 = " + str(count_mousemove_arrival))
if(count_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 0--->0 = " + str(count_mousemove_count))
if(count_mousemove_distance) != 0:exploredNodes.append("number of mousemove on distance, transition: 0--->0 = " + str(count_mousemove_distance))

if(count_mousemove_brush_departure) != 0:exploredNodes.append("number of special mousemove on departure, transition: 6--->6 = " + str(count_mousemove_brush_departure))
if(count_mousemove_brush_arrival) != 0:exploredNodes.append("number of special mousemove on arrival, transition: 5--->5 = " + str(count_mousemove_brush_arrival))
if(count_mousemove_brush_airtime) != 0:exploredNodes.append("number of special mousemove on airtime, transition: 7--->7 = " + str(count_mousemove_brush_airtime))
if(count_mousemove_brush_delay) != 0:exploredNodes.append("number of special mousemove on delay, transition: 8--->8 = " + str(count_mousemove_brush_delay))
if(count_mousemove_brush_distance) != 0:exploredNodes.append("number of special mousemove on distance, transition: 4--->4 = " + str(count_mousemove_brush_distance))

if(special_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 1--->1 = " + str(special_mousemove_count))

if(count_mousedown_distance) != 0:exploredNodes.append("number of mousedown on distance, transition: 0--->4 = " + str(count_mousedown_distance))
if(count_mousedown_departure) != 0:exploredNodes.append("number of mousedown on departure, transition: 0--->6 = " + str(count_mousedown_departure))
if(count_mousedown_arrival) != 0:exploredNodes.append("number of mousedown on arrival, transition: 0--->5 = " + str(count_mousedown_arrival))
if(count_mousedown_airtime) != 0:exploredNodes.append("number of mousedown on airtime, transition: 0--->7 = " + str(count_mousedown_airtime))
if(count_mousedown_delay) != 0:exploredNodes.append("number of mousedown on delay, transition: 0--->8 = " + str(count_mousedown_delay))
if(count_mousedown_count) != 0:exploredNodes.append("number of mousedown on count, transition: 0--->1 = " + str(count_mousedown_count))

if(count_mouseup_distance) != 0:exploredNodes.append("number of mouseup on distance, transition: 4--->0 = " + str(count_mouseup_distance))
if(count_mouseup_departure) != 0:exploredNodes.append("number of mouseup on departure, transition: 6--->0 = " + str(count_mouseup_departure))
if(count_mouseup_arrival) != 0:exploredNodes.append("number of mouseup on arrival, transition: 5--->0 = " + str(count_mouseup_arrival))
if(count_mouseup_airtime) != 0:exploredNodes.append("number of mouseup on airtime, transition: 7--->0 = " + str(count_mouseup_airtime))
if(count_mouseup_delay) != 0:exploredNodes.append("number of mouseup on delay, transition: 8--->0 = " + str(count_mouseup_delay))

if(count_click_distance) != 0:exploredNodes.append("number of click on distance, transition: 0--->0 = " + str(count_click_distance))
if(count_click_departure) != 0:exploredNodes.append("number of click on departure, transition: 0--->0 = " + str(count_click_departure))
if(count_click_arrival) != 0:exploredNodes.append("number of click on arrival, transition: 0--->0 = " + str(count_click_arrival))
if(count_click_airtime) != 0:exploredNodes.append("number of click on airtime, transition: 0--->0 = " + str(count_click_airtime))
if(count_click_delay) != 0:exploredNodes.append("number of click on delay, transition: 0--->0 = " + str(count_click_delay))
if(count_click_count) != 0:exploredNodes.append("number of click on count, transition: 0--->0 = " + str(count_click_count))
if(special_click_count) != 0:exploredNodes.append("number of special click on count, transition: 1--->1 = " + str(special_click_count))

if(count_dblclick_distance) != 0:exploredNodes.append("number of dblclick on distance, transition: 0--->0 = " + str(count_dblclick_distance))
if(count_dblclick_departure) != 0:exploredNodes.append("number of dblclick on departure, transition: 0--->0 = " + str(count_dblclick_departure))
if(count_dblclick_arrival) != 0:exploredNodes.append("number of dblclick on arrival, transition: 0--->0 = " + str(count_dblclick_arrival))
if(count_dblclick_airtime) != 0:exploredNodes.append("number of dblclick on airtime, transition: 0--->0 = " + str(count_dblclick_airtime))
if(count_dblclick_delay) != 0:exploredNodes.append("number of dblclick on delay, transition: 0--->0 = " + str(count_dblclick_delay))
if(count_dblclick_count) != 0:exploredNodes.append("number of dblclick on count, transition: 0--->0 = " + str(count_dblclick_count))

if(count_wheel_distance) != 0:exploredNodes.append("number of wheel on distance, transition: 0--->0 = " + str(count_wheel_distance))
if(count_wheel_departure) != 0:exploredNodes.append("number of wheel on departure, transition: 0--->0 = " + str(count_wheel_departure))
if(count_wheel_arrival) != 0:exploredNodes.append("number of wheel on arrival, transition: 0---> = " + str(count_wheel_arrival))
if(count_wheel_airtime) != 0:exploredNodes.append("number of wheel on airtime, transition: 0--->0 = " + str(count_wheel_airtime))
if(count_wheel_delay) != 0:exploredNodes.append("number of wheel on delay, transition: 0--->0 = " + str(count_wheel_delay))

if(special_wheel_distance) != 0:exploredNodes.append("number of special wheel on distance, transition: 4--->4 = " + str(special_wheel_distance))
if(special_wheel_departure) != 0:exploredNodes.append("number of special wheel on departure, transition: 6--->6 = " + str(special_wheel_departure))
if(special_wheel_arrival) != 0:exploredNodes.append("number of special wheel on arrival, transition: 5--->5 = " + str(special_wheel_arrival))
if(special_wheel_airtime) != 0:exploredNodes.append("number of special wheel on airtime, transition: 7--->7 = " + str(special_wheel_airtime))
if(special_wheel_delay) != 0:exploredNodes.append("number of special wheel on delay, transition: 8--->8 = " + str(special_wheel_delay))

if(count_mouseout_distance) != 0:exploredNodes.append("number of mouseout on distance, transition: 0--->0 = " + str(count_mouseout_distance))
if(count_mouseout_departure) != 0:exploredNodes.append("number of mouseout on departure, transition: 0--->0 = " + str(count_mouseout_departure))
if(count_mouseout_arrival) != 0:exploredNodes.append("number of mouseout on arrival, transition: 0--->0 = " + str(count_mouseout_arrival))
if(count_mouseout_airtime) != 0:exploredNodes.append("number of mouseout on airtime, transition: 0--->0 = " + str(count_mouseout_airtime))
if(count_mouseout_delay) != 0:exploredNodes.append("number of mouseout on delay, transition: 0--->0 = " + str(count_mouseout_delay))

if(special_mouseout_distance) != 0:exploredNodes.append("number of special mouseout on distance, transition: 4--->4 = " + str(special_mouseout_distance))
if(special_mouseout_departure) != 0:exploredNodes.append("number of special mouseout on departure, transition: 6--->6 = " + str(special_mouseout_departure))
if(special_mouseout_arrival) != 0:exploredNodes.append("number of special mouseout on arrival, transition: 5--->5 = " + str(special_mouseout_arrival))
if(special_mouseout_airtime) != 0:exploredNodes.append("number of special mouseout on airtime, transition: 7--->7 = " + str(special_mouseout_airtime))
if(special_mouseout_delay) != 0:exploredNodes.append("number of special mouseout on delay, transition: 8--->8 = " + str(special_mouseout_delay))
if(special_mouseout_count) != 0:exploredNodes.append("number of special mouseout on count, transition: 1--->1 = " + str(special_mouseout_count))

#save on excel file

sheet_total["A3"].value += count_mousemove_departure
sheet_total["B3"].value += count_mousemove_distance
sheet_total["C3"].value += count_mousemove_delay
sheet_total["D3"].value += count_mousemove_airtime
sheet_total["E3"].value += count_mousemove_arrival
sheet_total["F3"].value += count_mousemove_count
sheet_total["G3"].value += special_mousemove_count

sheet_total["A6"].value += count_mousedown_departure
sheet_total["B6"].value += count_mousedown_distance
sheet_total["C6"].value += count_mousedown_delay
sheet_total["D6"].value += count_mousedown_airtime
sheet_total["E6"].value += count_mousedown_arrival
sheet_total["F6"].value += count_mousedown_count

sheet_total["A10"].value += count_mousemove_brush_departure
sheet_total["B10"].value += count_mousemove_brush_distance
sheet_total["C10"].value += count_mousemove_brush_delay
sheet_total["D10"].value += count_mousemove_brush_airtime
sheet_total["E10"].value += count_mousemove_brush_arrival

sheet_total["A13"].value += count_mouseup_departure
sheet_total["B13"].value += count_mouseup_distance
sheet_total["C13"].value += count_mouseup_delay
sheet_total["D13"].value += count_mouseup_airtime
sheet_total["E13"].value += count_mouseup_arrival

sheet_total["A17"].value += count_click_departure
sheet_total["B17"].value += count_click_distance
sheet_total["C17"].value += count_click_delay
sheet_total["D17"].value += count_click_airtime
sheet_total["E17"].value += count_click_arrival
sheet_total["F17"].value += count_click_count
sheet_total["G17"].value += special_click_count

sheet_total["A21"].value += count_mouseout_departure
sheet_total["B21"].value += count_mouseout_distance
sheet_total["C21"].value += count_mouseout_delay
sheet_total["D21"].value += count_mouseout_airtime
sheet_total["E21"].value += count_mouseout_arrival
sheet_total["F21"].value += count_mouseout_count

sheet_total["A25"].value += special_mouseout_departure
sheet_total["B25"].value += special_mouseout_distance
sheet_total["C25"].value += special_mouseout_delay
sheet_total["D25"].value += special_mouseout_airtime
sheet_total["E25"].value += special_mouseout_arrival
sheet_total["F25"].value += special_mouseout_count

sheet_total["A29"].value += count_wheel_departure
sheet_total["B29"].value += count_wheel_distance
sheet_total["C29"].value += count_wheel_delay
sheet_total["D29"].value += count_wheel_airtime
sheet_total["E29"].value += count_wheel_arrival


sheet_total["A33"].value += special_wheel_departure
sheet_total["B33"].value += special_wheel_distance
sheet_total["C33"].value += special_wheel_delay
sheet_total["D33"].value += special_wheel_airtime
sheet_total["E33"].value += special_wheel_arrival

sheet_total["A37"].value += count_dblclick_departure
sheet_total["B37"].value += count_dblclick_distance
sheet_total["C37"].value += count_dblclick_delay
sheet_total["D37"].value += count_dblclick_airtime
sheet_total["E37"].value += count_dblclick_arrival
sheet_total["F37"].value += count_dblclick_count

wb_total.save(filename="total_count.xlsx")

sheet_4["A3"].value += count_mousemove_departure
sheet_4["B3"].value += count_mousemove_distance
sheet_4["C3"].value += count_mousemove_delay
sheet_4["D3"].value += count_mousemove_airtime
sheet_4["E3"].value += count_mousemove_arrival
sheet_4["F3"].value += count_mousemove_count
sheet_4["G3"].value += special_mousemove_count

sheet_4["A6"].value += count_mousedown_departure
sheet_4["B6"].value += count_mousedown_distance
sheet_4["C6"].value += count_mousedown_delay
sheet_4["D6"].value += count_mousedown_airtime
sheet_4["E6"].value += count_mousedown_arrival
sheet_4["F6"].value += count_mousedown_count

sheet_4["A10"].value += count_mousemove_brush_departure
sheet_4["B10"].value += count_mousemove_brush_distance
sheet_4["C10"].value += count_mousemove_brush_delay
sheet_4["D10"].value += count_mousemove_brush_airtime
sheet_4["E10"].value += count_mousemove_brush_arrival

sheet_4["A13"].value += count_mouseup_departure
sheet_4["B13"].value += count_mouseup_distance
sheet_4["C13"].value += count_mouseup_delay
sheet_4["D13"].value += count_mouseup_airtime
sheet_4["E13"].value += count_mouseup_arrival

sheet_4["A17"].value += count_click_departure
sheet_4["B17"].value += count_click_distance
sheet_4["C17"].value += count_click_delay
sheet_4["D17"].value += count_click_airtime
sheet_4["E17"].value += count_click_arrival
sheet_4["F17"].value += count_click_count
sheet_4["G17"].value += special_click_count

sheet_4["A21"].value += count_mouseout_departure
sheet_4["B21"].value += count_mouseout_distance
sheet_4["C21"].value += count_mouseout_delay
sheet_4["D21"].value += count_mouseout_airtime
sheet_4["E21"].value += count_mouseout_arrival
sheet_4["F21"].value += count_mouseout_count

sheet_4["A25"].value += special_mouseout_departure
sheet_4["B25"].value += special_mouseout_distance
sheet_4["C25"].value += special_mouseout_delay
sheet_4["D25"].value += special_mouseout_airtime
sheet_4["E25"].value += special_mouseout_arrival
sheet_4["F25"].value += special_mouseout_count

sheet_4["A29"].value += count_wheel_departure
sheet_4["B29"].value += count_wheel_distance
sheet_4["C29"].value += count_wheel_delay
sheet_4["D29"].value += count_wheel_airtime
sheet_4["E29"].value += count_wheel_arrival


sheet_4["A33"].value += special_wheel_departure
sheet_4["B33"].value += special_wheel_distance
sheet_4["C33"].value += special_wheel_delay
sheet_4["D33"].value += special_wheel_airtime
sheet_4["E33"].value += special_wheel_arrival

sheet_4["A37"].value += count_dblclick_departure
sheet_4["B37"].value += count_dblclick_distance
sheet_4["C37"].value += count_dblclick_delay
sheet_4["D37"].value += count_dblclick_airtime
sheet_4["E37"].value += count_dblclick_arrival
sheet_4["F37"].value += count_dblclick_count

wb_4.save(filename="count_path_4.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

###path 5!! da fare

for i in data["4"]:

    ### mousemove
    
    if(i == "mousemove on #distance canvas.marks , transition: 0--->0"):
        count_mousemove_distance = count_mousemove_distance + 1
    
    if(i == "mousemove on #arrival canvas.marks , transition: 0--->0"):
        count_mousemove_arrival = count_mousemove_arrival + 1
    
    if(i == "mousemove on #count canvas.marks , transition: 0--->0"):
        count_mousemove_count = count_mousemove_count + 1
    
    if(i == "mousemove on #departure canvas.marks , transition: 0--->0"):
        count_mousemove_departure = count_mousemove_departure + 1
    
    if(i == "mousemove on #airtime canvas.marks , transition: 0--->0"):
        count_mousemove_airtime = count_mousemove_airtime + 1
    
    if(i == "mousemove on #delay canvas.marks , transition: 0--->0"):
        count_mousemove_delay = count_mousemove_delay + 1
        
        ### mousemove brushing

    if(i == "mousemove on departure, transition: 6--->6"):
        count_mousemove_brush_departure +=1
    
    if(i == "mousemove on arrival, transition: 5--->5"):
        count_mousemove_brush_arrival +=1
    
    if(i == "mousemove on airtime, transition: 7--->7"):
        count_mousemove_brush_airtime +=1
    
    if(i == "mousemove on delay, transition: 8--->8"):
        count_mousemove_brush_delay +=1
    
    if(i == "mousemove on distance, transition: 4--->4"):
        count_mousemove_brush_distance +=1

    ### mousemove count

    if(i == "mousemove on count, transition: 1--->1"):
        special_mousemove_count+=1

    ### mousedown

    if(i == "mousedown on distance, transition: 0--->4"):
        count_mousedown_distance+=1

    if(i == "mousedown on departure, transition: 0--->6"):
        count_mousedown_departure+=1

    if(i == "mousedown on arrival, transition: 0--->5"):
        count_mousedown_arrival+=1

    if(i == "mousedown on airtime, transition: 0--->7"):
        count_mousedown_airtime+=1

    if(i == "mousedown on delay, transition: 0--->8"):
        count_mousedown_delay +=1

    if(i == "mousedown on count, transition: 0--->1"):
        count_mousedown_count +=1
    
    ### mouseup

    if(i == "mouseup on distance, transition: 4--->0"):
        count_mouseup_distance+=1

    if(i == "mouseup on departure, transition: 6--->0"):
        count_mouseup_departure+=1

    if(i == "mouseup on arrival, transition: 5--->0"):
        count_mouseup_arrival+=1

    if(i == "mouseup on airtime, transition: 7--->0"):
        count_mouseup_airtime+=1

    if(i == "mouseup on delay, transition: 8--->0"):
        count_mouseup_delay +=1

    ### click

    if(i == "click on #departure canvas.marks , transition: 0--->0"):
        count_click_departure+=1

    if(i == "click on #distance canvas.marks , transition: 0--->0"):
        count_click_distance+=1

    if(i == "click on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "click on #arrival canvas.marks , transition: 0--->0"):
        count_click_arrival+=1

    if(i == "click on #delay canvas.marks , transition: 0--->0"):
        count_click_delay +=1

    if(i == "click on #count canvas.marks , transition: 0--->0"):
        count_click_count +=1

    if(i == "click on #count canvas.marks , transition: 1--->1"):
        special_click_count +=1

    ### dblclick


    if(i == "dblclick on #departure canvas.marks , transition: 0--->0"):
        count_dblclick_departure+=1

    if(i == "dblclickk on #distance canvas.marks , transition: 0--->0"):
        count_dblclick_distance+=1

    if(i == "dblclick on #airtime canvas.marks , transition: 0--->0"):
        count_click_airtime+=1

    if(i == "dblclick on #arrival canvas.marks , transition: 0--->0"):
        count_dblclick_arrival+=1

    if(i == "dblclick on #delay canvas.marks , transition: 0--->0"):
        count_dblclick_delay +=1

    if(i == "dblclick on #count canvas.marks , transition: 0--->0"):
        count_dblclick_count +=1

    ### wheel

    if(i == "wheel on #distance canvas.marks , transition: 0--->0"):
        count_wheel_distance +=1
    
    if(i == "wheel on #departure canvas.marks , transition: 0--->0"):
        count_wheel_departure +=1
    
    if(i == "wheel on #airtime canvas.marks , transition: 0--->0"):
        count_wheel_airtime +=1
    
    if(i == "wheel on #arrival canvas.marks , transition: 0--->0"):
        count_wheel_arrival +=1
    
    if(i == "wheel on #delay canvas.marks , transition: 0--->0"):
        count_wheel_delay +=1

    ### special wheel

    if(i == "wheel on departure while brushing, transition: 6--->6"):
        special_wheel_departure +=1
    
    if(i == "wheel on arrival while brushing, transition: 5--->5"):
        special_wheel_arrival +=1
    
    if(i == "wheel on airtime while brushing, transition: 7--->7"):
        special_wheel_airtime +=1
    
    if(i == "wheel on delay while brushing, transition: 8--->8"):
        special_wheel_delay +=1
    
    if(i == "wheel on distance while brushing, transition: 4--->4"):
        special_wheel_distance +=1

    ### mouseout

    if(i == "mouseout on #distance canvas.marks , transition: 0--->0"):
        count_mouseout_distance +=1
    
    if(i == "mouseout on #departure canvas.marks , transition: 0--->0"):
        count_mouseout_departure +=1
    
    if(i == "mouseout on #airtime canvas.marks , transition: 0--->0"):
        count_mouseout_airtime +=1
    
    if(i == "mouseout on #arrival canvas.marks , transition: 0--->0"):
        count_mouseout_arrival +=1
    
    if(i == "mouseout on #delay canvas.marks , transition: 0--->0"):
        count_mouseout_delay +=1
    
    if(i == "mouseout on #count canvas.marks , transition: 0--->0"):
        count_mouseout_count +=1

    if(i == "mouseout on #count canvas.marks , transition: 1--->1"):
        special_mouseout_count +=1

    ### special mouseout

    if(i == "mouseout on departure while brushing, transition: 6--->6"):
        special_mouseout_departure +=1
    
    if(i == "mouseout on arrival while brushing, transition: 5--->5"):
        special_mouseout_arrival +=1
    
    if(i == "mouseout on airtime while brushing, transition: 7--->7"):
        special_mouseout_airtime +=1
    
    if(i == "mouseout on delay while brushing, transition: 8--->8"):
        special_mouseout_delay +=1
    
    if(i == "mouseout on distance while brushing, transition: 4--->4"):
        special_mouseout_distance +=1



exploredNodes.append("----------------------PATH 5------------------")

if(count_mousemove_departure) != 0:exploredNodes.append("number of mousemove on departure, transition: 0--->0 = " + str(count_mousemove_departure))
if(count_mousemove_delay) != 0:exploredNodes.append("number of mousemove on delay, transition: 0--->0 = " + str(count_mousemove_delay))
if(count_mousemove_airtime) != 0:exploredNodes.append("number of mousemove on airtime, transition: 0--->0 = " + str(count_mousemove_airtime))
if(count_mousemove_arrival) != 0:exploredNodes.append("number of mousemove on arrival, transition: 0--->0 = " + str(count_mousemove_arrival))
if(count_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 0--->0 = " + str(count_mousemove_count))
if(count_mousemove_distance) != 0:exploredNodes.append("number of mousemove on distance, transition: 0--->0 = " + str(count_mousemove_distance))

if(count_mousemove_brush_departure) != 0:exploredNodes.append("number of special mousemove on departure, transition: 6--->6 = " + str(count_mousemove_brush_departure))
if(count_mousemove_brush_arrival) != 0:exploredNodes.append("number of special mousemove on arrival, transition: 5--->5 = " + str(count_mousemove_brush_arrival))
if(count_mousemove_brush_airtime) != 0:exploredNodes.append("number of special mousemove on airtime, transition: 7--->7 = " + str(count_mousemove_brush_airtime))
if(count_mousemove_brush_delay) != 0:exploredNodes.append("number of special mousemove on delay, transition: 8--->8 = " + str(count_mousemove_brush_delay))
if(count_mousemove_brush_distance) != 0:exploredNodes.append("number of special mousemove on distance, transition: 4--->4 = " + str(count_mousemove_brush_distance))

if(special_mousemove_count) != 0:exploredNodes.append("number of mousemove on count, transition: 1--->1 = " + str(special_mousemove_count))

if(count_mousedown_distance) != 0:exploredNodes.append("number of mousedown on distance, transition: 0--->4 = " + str(count_mousedown_distance))
if(count_mousedown_departure) != 0:exploredNodes.append("number of mousedown on departure, transition: 0--->6 = " + str(count_mousedown_departure))
if(count_mousedown_arrival) != 0:exploredNodes.append("number of mousedown on arrival, transition: 0--->5 = " + str(count_mousedown_arrival))
if(count_mousedown_airtime) != 0:exploredNodes.append("number of mousedown on airtime, transition: 0--->7 = " + str(count_mousedown_airtime))
if(count_mousedown_delay) != 0:exploredNodes.append("number of mousedown on delay, transition: 0--->8 = " + str(count_mousedown_delay))
if(count_mousedown_count) != 0:exploredNodes.append("number of mousedown on count, transition: 0--->1 = " + str(count_mousedown_count))

if(count_mouseup_distance) != 0:exploredNodes.append("number of mouseup on distance, transition: 4--->0 = " + str(count_mouseup_distance))
if(count_mouseup_departure) != 0:exploredNodes.append("number of mouseup on departure, transition: 6--->0 = " + str(count_mouseup_departure))
if(count_mouseup_arrival) != 0:exploredNodes.append("number of mouseup on arrival, transition: 5--->0 = " + str(count_mouseup_arrival))
if(count_mouseup_airtime) != 0:exploredNodes.append("number of mouseup on airtime, transition: 7--->0 = " + str(count_mouseup_airtime))
if(count_mouseup_delay) != 0:exploredNodes.append("number of mouseup on delay, transition: 8--->0 = " + str(count_mouseup_delay))

if(count_click_distance) != 0:exploredNodes.append("number of click on distance, transition: 0--->0 = " + str(count_click_distance))
if(count_click_departure) != 0:exploredNodes.append("number of click on departure, transition: 0--->0 = " + str(count_click_departure))
if(count_click_arrival) != 0:exploredNodes.append("number of click on arrival, transition: 0--->0 = " + str(count_click_arrival))
if(count_click_airtime) != 0:exploredNodes.append("number of click on airtime, transition: 0--->0 = " + str(count_click_airtime))
if(count_click_delay) != 0:exploredNodes.append("number of click on delay, transition: 0--->0 = " + str(count_click_delay))
if(count_click_count) != 0:exploredNodes.append("number of click on count, transition: 0--->0 = " + str(count_click_count))
if(special_click_count) != 0:exploredNodes.append("number of special click on count, transition: 1--->1 = " + str(special_click_count))

if(count_dblclick_distance) != 0:exploredNodes.append("number of dblclick on distance, transition: 0--->0 = " + str(count_dblclick_distance))
if(count_dblclick_departure) != 0:exploredNodes.append("number of dblclick on departure, transition: 0--->0 = " + str(count_dblclick_departure))
if(count_dblclick_arrival) != 0:exploredNodes.append("number of dblclick on arrival, transition: 0--->0 = " + str(count_dblclick_arrival))
if(count_dblclick_airtime) != 0:exploredNodes.append("number of dblclick on airtime, transition: 0--->0 = " + str(count_dblclick_airtime))
if(count_dblclick_delay) != 0:exploredNodes.append("number of dblclick on delay, transition: 0--->0 = " + str(count_dblclick_delay))
if(count_dblclick_count) != 0:exploredNodes.append("number of dblclick on count, transition: 0--->0 = " + str(count_dblclick_count))

if(count_wheel_distance) != 0:exploredNodes.append("number of wheel on distance, transition: 0--->0 = " + str(count_wheel_distance))
if(count_wheel_departure) != 0:exploredNodes.append("number of wheel on departure, transition: 0--->0 = " + str(count_wheel_departure))
if(count_wheel_arrival) != 0:exploredNodes.append("number of wheel on arrival, transition: 0---> = " + str(count_wheel_arrival))
if(count_wheel_airtime) != 0:exploredNodes.append("number of wheel on airtime, transition: 0--->0 = " + str(count_wheel_airtime))
if(count_wheel_delay) != 0:exploredNodes.append("number of wheel on delay, transition: 0--->0 = " + str(count_wheel_delay))

if(special_wheel_distance) != 0:exploredNodes.append("number of special wheel on distance, transition: 4--->4 = " + str(special_wheel_distance))
if(special_wheel_departure) != 0:exploredNodes.append("number of special wheel on departure, transition: 6--->6 = " + str(special_wheel_departure))
if(special_wheel_arrival) != 0:exploredNodes.append("number of special wheel on arrival, transition: 5--->5 = " + str(special_wheel_arrival))
if(special_wheel_airtime) != 0:exploredNodes.append("number of special wheel on airtime, transition: 7--->7 = " + str(special_wheel_airtime))
if(special_wheel_delay) != 0:exploredNodes.append("number of special wheel on delay, transition: 8--->8 = " + str(special_wheel_delay))

if(count_mouseout_distance) != 0:exploredNodes.append("number of mouseout on distance, transition: 0--->0 = " + str(count_mouseout_distance))
if(count_mouseout_departure) != 0:exploredNodes.append("number of mouseout on departure, transition: 0--->0 = " + str(count_mouseout_departure))
if(count_mouseout_arrival) != 0:exploredNodes.append("number of mouseout on arrival, transition: 0--->0 = " + str(count_mouseout_arrival))
if(count_mouseout_airtime) != 0:exploredNodes.append("number of mouseout on airtime, transition: 0--->0 = " + str(count_mouseout_airtime))
if(count_mouseout_delay) != 0:exploredNodes.append("number of mouseout on delay, transition: 0--->0 = " + str(count_mouseout_delay))

if(special_mouseout_distance) != 0:exploredNodes.append("number of special mouseout on distance, transition: 4--->4 = " + str(special_mouseout_distance))
if(special_mouseout_departure) != 0:exploredNodes.append("number of special mouseout on departure, transition: 6--->6 = " + str(special_mouseout_departure))
if(special_mouseout_arrival) != 0:exploredNodes.append("number of special mouseout on arrival, transition: 5--->5 = " + str(special_mouseout_arrival))
if(special_mouseout_airtime) != 0:exploredNodes.append("number of special mouseout on airtime, transition: 7--->7 = " + str(special_mouseout_airtime))
if(special_mouseout_delay) != 0:exploredNodes.append("number of special mouseout on delay, transition: 8--->8 = " + str(special_mouseout_delay))
if(special_mouseout_count) != 0:exploredNodes.append("number of special mouseout on count, transition: 1--->1 = " + str(special_mouseout_count))

#save on excel file

sheet_total["A3"].value += count_mousemove_departure
sheet_total["B3"].value += count_mousemove_distance
sheet_total["C3"].value += count_mousemove_delay
sheet_total["D3"].value += count_mousemove_airtime
sheet_total["E3"].value += count_mousemove_arrival
sheet_total["F3"].value += count_mousemove_count
sheet_total["G3"].value += special_mousemove_count

sheet_total["A6"].value += count_mousedown_departure
sheet_total["B6"].value += count_mousedown_distance
sheet_total["C6"].value += count_mousedown_delay
sheet_total["D6"].value += count_mousedown_airtime
sheet_total["E6"].value += count_mousedown_arrival
sheet_total["F6"].value += count_mousedown_count

sheet_total["A10"].value += count_mousemove_brush_departure
sheet_total["B10"].value += count_mousemove_brush_distance
sheet_total["C10"].value += count_mousemove_brush_delay
sheet_total["D10"].value += count_mousemove_brush_airtime
sheet_total["E10"].value += count_mousemove_brush_arrival

sheet_total["A13"].value += count_mouseup_departure
sheet_total["B13"].value += count_mouseup_distance
sheet_total["C13"].value += count_mouseup_delay
sheet_total["D13"].value += count_mouseup_airtime
sheet_total["E13"].value += count_mouseup_arrival

sheet_total["A17"].value += count_click_departure
sheet_total["B17"].value += count_click_distance
sheet_total["C17"].value += count_click_delay
sheet_total["D17"].value += count_click_airtime
sheet_total["E17"].value += count_click_arrival
sheet_total["F17"].value += count_click_count
sheet_total["G17"].value += special_click_count

sheet_total["A21"].value += count_mouseout_departure
sheet_total["B21"].value += count_mouseout_distance
sheet_total["C21"].value += count_mouseout_delay
sheet_total["D21"].value += count_mouseout_airtime
sheet_total["E21"].value += count_mouseout_arrival
sheet_total["F21"].value += count_mouseout_count

sheet_total["A25"].value += special_mouseout_departure
sheet_total["B25"].value += special_mouseout_distance
sheet_total["C25"].value += special_mouseout_delay
sheet_total["D25"].value += special_mouseout_airtime
sheet_total["E25"].value += special_mouseout_arrival
sheet_total["F25"].value += special_mouseout_count

sheet_total["A29"].value += count_wheel_departure
sheet_total["B29"].value += count_wheel_distance
sheet_total["C29"].value += count_wheel_delay
sheet_total["D29"].value += count_wheel_airtime
sheet_total["E29"].value += count_wheel_arrival


sheet_total["A33"].value += special_wheel_departure
sheet_total["B33"].value += special_wheel_distance
sheet_total["C33"].value += special_wheel_delay
sheet_total["D33"].value += special_wheel_airtime
sheet_total["E33"].value += special_wheel_arrival

sheet_total["A37"].value += count_dblclick_departure
sheet_total["B37"].value += count_dblclick_distance
sheet_total["C37"].value += count_dblclick_delay
sheet_total["D37"].value += count_dblclick_airtime
sheet_total["E37"].value += count_dblclick_arrival
sheet_total["F37"].value += count_dblclick_count

wb_total.save(filename="total_count.xlsx")

sheet_5["A3"].value += count_mousemove_departure
sheet_5["B3"].value += count_mousemove_distance
sheet_5["C3"].value += count_mousemove_delay
sheet_5["D3"].value += count_mousemove_airtime
sheet_5["E3"].value += count_mousemove_arrival
sheet_5["F3"].value += count_mousemove_count
sheet_5["G3"].value += special_mousemove_count

sheet_5["A6"].value += count_mousedown_departure
sheet_5["B6"].value += count_mousedown_distance
sheet_5["C6"].value += count_mousedown_delay
sheet_5["D6"].value += count_mousedown_airtime
sheet_5["E6"].value += count_mousedown_arrival
sheet_5["F6"].value += count_mousedown_count

sheet_5["A10"].value += count_mousemove_brush_departure
sheet_5["B10"].value += count_mousemove_brush_distance
sheet_5["C10"].value += count_mousemove_brush_delay
sheet_5["D10"].value += count_mousemove_brush_airtime
sheet_5["E10"].value += count_mousemove_brush_arrival

sheet_5["A13"].value += count_mouseup_departure
sheet_5["B13"].value += count_mouseup_distance
sheet_5["C13"].value += count_mouseup_delay
sheet_5["D13"].value += count_mouseup_airtime
sheet_5["E13"].value += count_mouseup_arrival

sheet_5["A17"].value += count_click_departure
sheet_5["B17"].value += count_click_distance
sheet_5["C17"].value += count_click_delay
sheet_5["D17"].value += count_click_airtime
sheet_5["E17"].value += count_click_arrival
sheet_5["F17"].value += count_click_count
sheet_5["G17"].value += special_click_count

sheet_5["A21"].value += count_mouseout_departure
sheet_5["B21"].value += count_mouseout_distance
sheet_5["C21"].value += count_mouseout_delay
sheet_5["D21"].value += count_mouseout_airtime
sheet_5["E21"].value += count_mouseout_arrival
sheet_5["F21"].value += count_mouseout_count

sheet_5["A25"].value += special_mouseout_departure
sheet_5["B25"].value += special_mouseout_distance
sheet_5["C25"].value += special_mouseout_delay
sheet_5["D25"].value += special_mouseout_airtime
sheet_5["E25"].value += special_mouseout_arrival
sheet_5["F25"].value += special_mouseout_count

sheet_5["A29"].value += count_wheel_departure
sheet_5["B29"].value += count_wheel_distance
sheet_5["C29"].value += count_wheel_delay
sheet_5["D29"].value += count_wheel_airtime
sheet_5["E29"].value += count_wheel_arrival


sheet_5["A33"].value += special_wheel_departure
sheet_5["B33"].value += special_wheel_distance
sheet_5["C33"].value += special_wheel_delay
sheet_5["D33"].value += special_wheel_airtime
sheet_5["E33"].value += special_wheel_arrival

sheet_5["A37"].value += count_dblclick_departure
sheet_5["B37"].value += count_dblclick_distance
sheet_5["C37"].value += count_dblclick_delay
sheet_5["D37"].value += count_dblclick_airtime
sheet_5["E37"].value += count_dblclick_arrival
sheet_5["F37"].value += count_dblclick_count

wb_5.save(filename="count_path_5.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dblclick

count_dblclick_delay = 0
count_dblclick_airtime = 0
count_dblclick_distance = 0
count_dblclick_arrival = 0
count_dblclick_departure = 0
count_dblclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0
    
        
    
  

# Closing file
with open('count_' + "falcon_7M" + '_50' + '.json', 'w') as fp:
            json.dump(exploredNodes, fp,  indent=4)

f.close()