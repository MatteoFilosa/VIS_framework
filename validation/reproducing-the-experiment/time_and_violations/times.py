import json
from openpyxl import Workbook
from openpyxl import load_workbook

#workbook = Workbook()
wb_total = load_workbook(filename = 'total_time.xlsx')
sheet_total = wb_total.active

wb_1 = load_workbook(filename = 'time_path_1.xlsx')
sheet_1 = wb_1.active


wb_2 = load_workbook(filename = 'time_path_2.xlsx')
sheet_2 = wb_2.active


wb_3 = load_workbook(filename = 'time_path_3.xlsx')
sheet_3 = wb_3.active


wb_4 = load_workbook(filename = 'time_path_4.xlsx')
sheet_4 = wb_4.active

wb_5 = load_workbook(filename = 'time_path_5.xlsx')
sheet_5 = wb_5.active

  
# Opening JSON file
f = open('summary_falcon_7M_1.json') #only opens one exploration file

#mousemove
count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

##########################TIME############################

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#remember!

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0

pathNumber = 0
violationsFound = {}
violationsFound[0] = []
violationsFound[1] = []
violationsFound[2] = []
violationsFound[3] = []
violationsFound[4] = []
# returns JSON object as 
# a dictionary
data = json.load(f)

for i in data["0"]:

    ### mousemove
    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_count += 1
        time_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_distance += 1
        time_mousemove_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_arrival += 1
        time_mousemove_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_departure += 1
        time_mousemove_departure += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_airtime += 1
        time_mousemove_airtime += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_delay += 1
        time_mousemove_delay += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

        ### mousemove brushing


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_distance += 1
        time_mousemove_brush_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_arrival += 1
        time_mousemove_brush_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_departure += 1
        time_mousemove_brush_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_airtime += 1
        time_mousemove_brush_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_delay += 1
        time_mousemove_brush_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    ### mousemove count, later!!! detector per la mousedown prima.

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mousemove_count += 1
        time_special_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mousedown

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_count += 1
        time_mousedown_count += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_distance += 1
        time_mousedown_distance += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_arrival += 1
        time_mousedown_arrival += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_departure += 1
        time_mousedown_departure += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_airtime += 1
        time_mousedown_airtime += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_delay += 1
        time_mousedown_delay += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mouseup


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_distance += 1
        time_mouseup_distance += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_arrival += 1
        time_mouseup_arrival += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_departure += 1
        time_mouseup_departure += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_airtime += 1
        time_mouseup_airtime += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_delay += 1
        time_mouseup_delay += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### click

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "click") :
        
        count_click_count += 1
        time_click_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "click") :
        
        count_click_distance += 1
        time_click_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "click") :
        
        count_click_arrival += 1
        time_click_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "click") :
        
        count_click_departure += 1
        time_click_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "click") :
        
        count_click_airtime += 1
        time_click_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "click") :
        
        count_click_delay += 1
        time_click_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### dbclick

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_count += 1
        time_dbclick_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_distance += 1
        time_dbclick_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_arrival += 1
        time_dbclick_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_departure += 1
        time_dbclick_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_airtime += 1
        time_dbclick_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_delay += 1
        time_dbclick_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### wheel


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_distance += 1
        time_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_arrival += 1
        time_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_departure += 1
        time_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_airtime += 1
        time_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_delay += 1
        time_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### special wheel 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_wheel_distance += 1
        time_special_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_arrival += 1
        time_special_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_departure += 1
        time_special_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_airtime += 1
        time_special_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_delay += 1
        time_special_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### mouseout

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_count += 1
        time_mouseout_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_distance += 1
        time_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_arrival += 1
        time_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_departure += 1
        time_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)  
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_airtime += 1
        time_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_delay += 1
        time_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### special mouseout 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mouseout_distance += 1
        time_special_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
            
    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_arrival += 1
        time_special_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_departure += 1
        time_special_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_airtime += 1
        time_special_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_delay += 1
        time_special_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


#save on excel file

if (count_mousemove_departure != 0) : sheet_total["A3"].value = (sheet_total["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_total["B3"].value = (sheet_total["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_total["C3"].value = (sheet_total["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_total["D3"].value = (sheet_total["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_total["E3"].value = (sheet_total["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_total["F3"].value = (sheet_total["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_total["G3"].value = (sheet_total["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_total["A6"].value = (sheet_total["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_total["B6"].value = (sheet_total["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_total["C6"].value = (sheet_total["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_total["D6"].value = (sheet_total["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_total["E6"].value = (sheet_total["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_total["F6"].value = (sheet_total["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_total["A10"].value = (sheet_total["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_total["B10"].value = (sheet_total["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_total["C10"].value = (sheet_total["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_total["D10"].value = (sheet_total["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_total["E10"].value = (sheet_total["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_total["A13"].value = (sheet_total["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_total["B13"].value = (sheet_total["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_total["C13"].value = (sheet_total["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_total["D13"].value = (sheet_total["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_total["E13"].value = (sheet_total["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_total["A17"].value = (sheet_total["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_total["B17"].value = (sheet_total["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_total["C17"].value = (sheet_total["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_total["D17"].value = (sheet_total["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_total["E17"].value = (sheet_total["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_total["F17"].value = (sheet_total["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_total["G17"].value = (sheet_total["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_total["A21"].value = (sheet_total["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_total["B21"].value = (sheet_total["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_total["C21"].value = (sheet_total["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_total["D21"].value = (sheet_total["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_total["E21"].value = (sheet_total["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_total["F21"].value = (sheet_total["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_total["A25"].value = (sheet_total["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_total["B25"].value = (sheet_total["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_total["C25"].value = (sheet_total["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_total["D25"].value = (sheet_total["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_total["E25"].value = (sheet_total["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_total["F25"].value = (sheet_total["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_total["A29"].value = (sheet_total["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_total["B29"].value = (sheet_total["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_total["C29"].value = (sheet_total["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_total["D29"].value = (sheet_total["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_total["E29"].value = (sheet_total["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_total["A33"].value = (sheet_total["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_total["B33"].value = (sheet_total["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_total["C33"].value = (sheet_total["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_total["D33"].value = (sheet_total["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_total["E33"].value = (sheet_total["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_total["A37"].value = (sheet_total["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_total["B37"].value = (sheet_total["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_total["C37"].value = (sheet_total["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_total["D37"].value = (sheet_total["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_total["E37"].value = (sheet_total["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_total["F37"].value = (sheet_total["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_total.save(filename="total_time.xlsx")

#saving to the relative path excel file


if (count_mousemove_departure != 0) : sheet_1["A3"].value = (sheet_1["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_1["B3"].value = (sheet_1["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_1["C3"].value = (sheet_1["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_1["D3"].value = (sheet_1["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_1["E3"].value = (sheet_1["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_1["F3"].value = (sheet_1["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_1["G3"].value = (sheet_1["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_1["A6"].value = (sheet_1["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_1["B6"].value = (sheet_1["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_1["C6"].value = (sheet_1["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_1["D6"].value = (sheet_1["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_1["E6"].value = (sheet_1["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_1["F6"].value = (sheet_1["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_1["A10"].value = (sheet_1["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_1["B10"].value = (sheet_1["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_1["C10"].value = (sheet_1["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_1["D10"].value = (sheet_1["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_1["E10"].value = (sheet_1["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_1["A13"].value = (sheet_1["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_1["B13"].value = (sheet_1["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_1["C13"].value = (sheet_1["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_1["D13"].value = (sheet_1["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_1["E13"].value = (sheet_1["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_1["A17"].value = (sheet_1["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_1["B17"].value = (sheet_1["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_1["C17"].value = (sheet_1["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_1["D17"].value = (sheet_1["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_1["E17"].value = (sheet_1["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_1["F17"].value = (sheet_1["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_1["G17"].value = (sheet_1["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_1["A21"].value = (sheet_1["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_1["B21"].value = (sheet_1["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_1["C21"].value = (sheet_1["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_1["D21"].value = (sheet_1["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_1["E21"].value = (sheet_1["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_1["F21"].value = (sheet_1["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_1["A25"].value = (sheet_1["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_1["B25"].value = (sheet_1["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_1["C25"].value = (sheet_1["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_1["D25"].value = (sheet_1["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_1["E25"].value = (sheet_1["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_1["F25"].value = (sheet_1["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_1["A29"].value = (sheet_1["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_1["B29"].value = (sheet_1["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_1["C29"].value = (sheet_1["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_1["D29"].value = (sheet_1["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_1["E29"].value = (sheet_1["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_1["A33"].value = (sheet_1["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_1["B33"].value = (sheet_1["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_1["C33"].value = (sheet_1["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_1["D33"].value = (sheet_1["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_1["E33"].value = (sheet_1["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_1["A37"].value = (sheet_1["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_1["B37"].value = (sheet_1["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_1["C37"].value = (sheet_1["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_1["D37"].value = (sheet_1["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_1["E37"].value = (sheet_1["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_1["F37"].value = (sheet_1["F37"].value + (time_dbclick_count/count_dbclick_count))/2


wb_1.save(filename="time_path_1.xlsx")

#setting back to 0 the variables

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

####also the time variables to 0!!!!

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#and the remember variables...

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0

for i in data["1"]:

    pathNumber = 1

    ### mousemove
    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_count += 1
        time_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_distance += 1
        time_mousemove_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_arrival += 1
        time_mousemove_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_departure += 1
        time_mousemove_departure += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_airtime += 1
        time_mousemove_airtime += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_delay += 1
        time_mousemove_delay += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

        ### mousemove brushing


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_distance += 1
        time_mousemove_brush_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_arrival += 1
        time_mousemove_brush_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_departure += 1
        time_mousemove_brush_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_airtime += 1
        time_mousemove_brush_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_delay += 1
        time_mousemove_brush_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    ### mousemove count, later!!! detector per la mousedown prima.

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mousemove_count += 1
        time_special_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mousedown

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_count += 1
        time_mousedown_count += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_distance += 1
        time_mousedown_distance += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_arrival += 1
        time_mousedown_arrival += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_departure += 1
        time_mousedown_departure += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_airtime += 1
        time_mousedown_airtime += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_delay += 1
        time_mousedown_delay += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mouseup


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_distance += 1
        time_mouseup_distance += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_arrival += 1
        time_mouseup_arrival += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_departure += 1
        time_mouseup_departure += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_airtime += 1
        time_mouseup_airtime += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_delay += 1
        time_mouseup_delay += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### click

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "click") :
        
        count_click_count += 1
        time_click_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "click") :
        
        count_click_distance += 1
        time_click_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "click") :
        
        count_click_arrival += 1
        time_click_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "click") :
        
        count_click_departure += 1
        time_click_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "click") :
        
        count_click_airtime += 1
        time_click_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "click") :
        
        count_click_delay += 1
        time_click_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### dbclick

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_count += 1
        time_dbclick_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_distance += 1
        time_dbclick_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_arrival += 1
        time_dbclick_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_departure += 1
        time_dbclick_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_airtime += 1
        time_dbclick_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_delay += 1
        time_dbclick_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### wheel


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_distance += 1
        time_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_arrival += 1
        time_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_departure += 1
        time_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_airtime += 1
        time_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_delay += 1
        time_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### special wheel 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_wheel_distance += 1
        time_special_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_arrival += 1
        time_special_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_departure += 1
        time_special_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_airtime += 1
        time_special_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_delay += 1
        time_special_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### mouseout

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_count += 1
        time_mouseout_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_distance += 1
        time_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_arrival += 1
        time_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_departure += 1
        time_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)  
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_airtime += 1
        time_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_delay += 1
        time_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### special mouseout 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mouseout_distance += 1
        time_special_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
            
    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_arrival += 1
        time_special_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_departure += 1
        time_special_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_airtime += 1
        time_special_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_delay += 1
        time_special_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

#save on excel file



if (count_mousemove_departure != 0) : sheet_total["A3"].value = (sheet_total["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_total["B3"].value = (sheet_total["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_total["C3"].value = (sheet_total["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_total["D3"].value = (sheet_total["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_total["E3"].value = (sheet_total["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_total["F3"].value = (sheet_total["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_total["G3"].value = (sheet_total["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_total["A6"].value = (sheet_total["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_total["B6"].value = (sheet_total["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_total["C6"].value = (sheet_total["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_total["D6"].value = (sheet_total["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_total["E6"].value = (sheet_total["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_total["F6"].value = (sheet_total["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_total["A10"].value = (sheet_total["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_total["B10"].value = (sheet_total["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_total["C10"].value = (sheet_total["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_total["D10"].value = (sheet_total["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_total["E10"].value = (sheet_total["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_total["A13"].value = (sheet_total["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_total["B13"].value = (sheet_total["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_total["C13"].value = (sheet_total["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_total["D13"].value = (sheet_total["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_total["E13"].value = (sheet_total["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_total["A17"].value = (sheet_total["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_total["B17"].value = (sheet_total["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_total["C17"].value = (sheet_total["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_total["D17"].value = (sheet_total["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_total["E17"].value = (sheet_total["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_total["F17"].value = (sheet_total["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_total["G17"].value = (sheet_total["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_total["A21"].value = (sheet_total["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_total["B21"].value = (sheet_total["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_total["C21"].value = (sheet_total["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_total["D21"].value = (sheet_total["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_total["E21"].value = (sheet_total["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_total["F21"].value = (sheet_total["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_total["A25"].value = (sheet_total["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_total["B25"].value = (sheet_total["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_total["C25"].value = (sheet_total["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_total["D25"].value = (sheet_total["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_total["E25"].value = (sheet_total["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_total["F25"].value = (sheet_total["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_total["A29"].value = (sheet_total["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_total["B29"].value = (sheet_total["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_total["C29"].value = (sheet_total["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_total["D29"].value = (sheet_total["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_total["E29"].value = (sheet_total["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_total["A33"].value = (sheet_total["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_total["B33"].value = (sheet_total["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_total["C33"].value = (sheet_total["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_total["D33"].value = (sheet_total["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_total["E33"].value = (sheet_total["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_total["A37"].value = (sheet_total["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_total["B37"].value = (sheet_total["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_total["C37"].value = (sheet_total["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_total["D37"].value = (sheet_total["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_total["E37"].value = (sheet_total["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_total["F37"].value = (sheet_total["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_total.save(filename="total_time.xlsx")


if (count_mousemove_departure != 0) : sheet_2["A3"].value = (sheet_2["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_2["B3"].value = (sheet_2["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_2["C3"].value = (sheet_2["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_2["D3"].value = (sheet_2["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_2["E3"].value = (sheet_2["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_2["F3"].value = (sheet_2["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_2["G3"].value = (sheet_2["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_2["A6"].value = (sheet_2["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_2["B6"].value = (sheet_2["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_2["C6"].value = (sheet_2["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_2["D6"].value = (sheet_2["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_2["E6"].value = (sheet_2["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_2["F6"].value = (sheet_2["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_2["A10"].value = (sheet_2["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_2["B10"].value = (sheet_2["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_2["C10"].value = (sheet_2["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_2["D10"].value = (sheet_2["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_2["E10"].value = (sheet_2["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_2["A13"].value = (sheet_2["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_2["B13"].value = (sheet_2["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_2["C13"].value = (sheet_2["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_2["D13"].value = (sheet_2["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_2["E13"].value = (sheet_2["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_2["A17"].value = (sheet_2["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_2["B17"].value = (sheet_2["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_2["C17"].value = (sheet_2["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_2["D17"].value = (sheet_2["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_2["E17"].value = (sheet_2["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_2["F17"].value = (sheet_2["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_2["G17"].value = (sheet_2["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_2["A21"].value = (sheet_2["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_2["B21"].value = (sheet_2["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_2["C21"].value = (sheet_2["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_2["D21"].value = (sheet_2["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_2["E21"].value = (sheet_2["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_2["F21"].value = (sheet_2["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_2["A25"].value = (sheet_2["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_2["B25"].value = (sheet_2["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_2["C25"].value = (sheet_2["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_2["D25"].value = (sheet_2["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_2["E25"].value = (sheet_2["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_2["F25"].value = (sheet_2["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_2["A29"].value = (sheet_2["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_2["B29"].value = (sheet_2["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_2["C29"].value = (sheet_2["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_2["D29"].value = (sheet_2["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_2["E29"].value = (sheet_2["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_2["A33"].value = (sheet_2["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_2["B33"].value = (sheet_2["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_2["C33"].value = (sheet_2["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_2["D33"].value = (sheet_2["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_2["E33"].value = (sheet_2["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_2["A37"].value = (sheet_2["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_2["B37"].value = (sheet_2["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_2["C37"].value = (sheet_2["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_2["D37"].value = (sheet_2["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_2["E37"].value = (sheet_2["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_2["F37"].value = (sheet_2["F37"].value + (time_dbclick_count/count_dbclick_count))/2


wb_2.save(filename="time_path_2.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

####also the time variables to 0!!!!

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#and the remember variables...

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0

for i in data["2"]:

    pathNumber = 2
    ### mousemove
    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_count += 1
        time_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_distance += 1
        time_mousemove_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_arrival += 1
        time_mousemove_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_departure += 1
        time_mousemove_departure += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_airtime += 1
        time_mousemove_airtime += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_delay += 1
        time_mousemove_delay += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

        ### mousemove brushing


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_distance += 1
        time_mousemove_brush_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_arrival += 1
        time_mousemove_brush_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_departure += 1
        time_mousemove_brush_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_airtime += 1
        time_mousemove_brush_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_delay += 1
        time_mousemove_brush_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    ### mousemove count, later!!! detector per la mousedown prima.

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mousemove_count += 1
        time_special_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mousedown

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_count += 1
        time_mousedown_count += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_distance += 1
        time_mousedown_distance += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_arrival += 1
        time_mousedown_arrival += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_departure += 1
        time_mousedown_departure += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_airtime += 1
        time_mousedown_airtime += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_delay += 1
        time_mousedown_delay += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mouseup


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_distance += 1
        time_mouseup_distance += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_arrival += 1
        time_mouseup_arrival += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_departure += 1
        time_mouseup_departure += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_airtime += 1
        time_mouseup_airtime += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_delay += 1
        time_mouseup_delay += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### click

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "click") :
        
        count_click_count += 1
        time_click_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "click") :
        
        count_click_distance += 1
        time_click_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "click") :
        
        count_click_arrival += 1
        time_click_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "click") :
        
        count_click_departure += 1
        time_click_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "click") :
        
        count_click_airtime += 1
        time_click_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "click") :
        
        count_click_delay += 1
        time_click_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### dbclick

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_count += 1
        time_dbclick_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_distance += 1
        time_dbclick_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_arrival += 1
        time_dbclick_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_departure += 1
        time_dbclick_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_airtime += 1
        time_dbclick_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_delay += 1
        time_dbclick_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### wheel


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_distance += 1
        time_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_arrival += 1
        time_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_departure += 1
        time_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_airtime += 1
        time_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_delay += 1
        time_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### special wheel 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_wheel_distance += 1
        time_special_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_arrival += 1
        time_special_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_departure += 1
        time_special_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_airtime += 1
        time_special_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_delay += 1
        time_special_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### mouseout

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_count += 1
        time_mouseout_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_distance += 1
        time_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_arrival += 1
        time_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_departure += 1
        time_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)  
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_airtime += 1
        time_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_delay += 1
        time_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### special mouseout 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mouseout_distance += 1
        time_special_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
            
    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_arrival += 1
        time_special_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_departure += 1
        time_special_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_airtime += 1
        time_special_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_delay += 1
        time_special_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


#save on excel file



if (count_mousemove_departure != 0) : sheet_total["A3"].value = (sheet_total["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_total["B3"].value = (sheet_total["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_total["C3"].value = (sheet_total["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_total["D3"].value = (sheet_total["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_total["E3"].value = (sheet_total["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_total["F3"].value = (sheet_total["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_total["G3"].value = (sheet_total["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_total["A6"].value = (sheet_total["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_total["B6"].value = (sheet_total["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_total["C6"].value = (sheet_total["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_total["D6"].value = (sheet_total["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_total["E6"].value = (sheet_total["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_total["F6"].value = (sheet_total["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_total["A10"].value = (sheet_total["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_total["B10"].value = (sheet_total["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_total["C10"].value = (sheet_total["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_total["D10"].value = (sheet_total["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_total["E10"].value = (sheet_total["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_total["A13"].value = (sheet_total["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_total["B13"].value = (sheet_total["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_total["C13"].value = (sheet_total["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_total["D13"].value = (sheet_total["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_total["E13"].value = (sheet_total["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_total["A17"].value = (sheet_total["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_total["B17"].value = (sheet_total["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_total["C17"].value = (sheet_total["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_total["D17"].value = (sheet_total["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_total["E17"].value = (sheet_total["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_total["F17"].value = (sheet_total["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_total["G17"].value = (sheet_total["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_total["A21"].value = (sheet_total["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_total["B21"].value = (sheet_total["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_total["C21"].value = (sheet_total["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_total["D21"].value = (sheet_total["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_total["E21"].value = (sheet_total["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_total["F21"].value = (sheet_total["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_total["A25"].value = (sheet_total["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_total["B25"].value = (sheet_total["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_total["C25"].value = (sheet_total["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_total["D25"].value = (sheet_total["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_total["E25"].value = (sheet_total["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_total["F25"].value = (sheet_total["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_total["A29"].value = (sheet_total["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_total["B29"].value = (sheet_total["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_total["C29"].value = (sheet_total["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_total["D29"].value = (sheet_total["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_total["E29"].value = (sheet_total["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_total["A33"].value = (sheet_total["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_total["B33"].value = (sheet_total["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_total["C33"].value = (sheet_total["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_total["D33"].value = (sheet_total["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_total["E33"].value = (sheet_total["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_total["A37"].value = (sheet_total["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_total["B37"].value = (sheet_total["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_total["C37"].value = (sheet_total["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_total["D37"].value = (sheet_total["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_total["E37"].value = (sheet_total["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_total["F37"].value = (sheet_total["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_total.save(filename="total_time.xlsx")


if (count_mousemove_departure != 0) : sheet_3["A3"].value = (sheet_3["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_3["B3"].value = (sheet_3["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_3["C3"].value = (sheet_3["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_3["D3"].value = (sheet_3["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_3["E3"].value = (sheet_3["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_3["F3"].value = (sheet_3["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_3["G3"].value = (sheet_3["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_3["A6"].value = (sheet_3["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_3["B6"].value = (sheet_3["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_3["C6"].value = (sheet_3["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_3["D6"].value = (sheet_3["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_3["E6"].value = (sheet_3["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_3["F6"].value = (sheet_3["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_3["A10"].value = (sheet_3["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_3["B10"].value = (sheet_3["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_3["C10"].value = (sheet_3["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_3["D10"].value = (sheet_3["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_3["E10"].value = (sheet_3["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_3["A13"].value = (sheet_3["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_3["B13"].value = (sheet_3["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_3["C13"].value = (sheet_3["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_3["D13"].value = (sheet_3["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_3["E13"].value = (sheet_3["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_3["A17"].value = (sheet_3["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_3["B17"].value = (sheet_3["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_3["C17"].value = (sheet_3["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_3["D17"].value = (sheet_3["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_3["E17"].value = (sheet_3["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_3["F17"].value = (sheet_3["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_3["G17"].value = (sheet_3["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_3["A21"].value = (sheet_3["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_3["B21"].value = (sheet_3["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_3["C21"].value = (sheet_3["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_3["D21"].value = (sheet_3["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_3["E21"].value = (sheet_3["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_3["F21"].value = (sheet_3["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_3["A25"].value = (sheet_3["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_3["B25"].value = (sheet_3["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_3["C25"].value = (sheet_3["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_3["D25"].value = (sheet_3["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_3["E25"].value = (sheet_3["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_3["F25"].value = (sheet_3["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_3["A29"].value = (sheet_3["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_3["B29"].value = (sheet_3["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_3["C29"].value = (sheet_3["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_3["D29"].value = (sheet_3["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_3["E29"].value = (sheet_3["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_3["A33"].value = (sheet_3["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_3["B33"].value = (sheet_3["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_3["C33"].value = (sheet_3["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_3["D33"].value = (sheet_3["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_3["E33"].value = (sheet_3["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_3["A37"].value = (sheet_3["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_3["B37"].value = (sheet_3["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_3["C37"].value = (sheet_3["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_3["D37"].value = (sheet_3["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_3["E37"].value = (sheet_3["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_3["F37"].value = (sheet_3["F37"].value + (time_dbclick_count/count_dbclick_count))/2


wb_3.save(filename="time_path_3.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0
####also the time variables to 0!!!!

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#and the remember variables...

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0

for i in data["3"]:

    pathNumber = 3

    ### mousemove
    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_count += 1
        time_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_distance += 1
        time_mousemove_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_arrival += 1
        time_mousemove_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_departure += 1
        time_mousemove_departure += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_airtime += 1
        time_mousemove_airtime += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_delay += 1
        time_mousemove_delay += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

        ### mousemove brushing


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_distance += 1
        time_mousemove_brush_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_arrival += 1
        time_mousemove_brush_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_departure += 1
        time_mousemove_brush_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_airtime += 1
        time_mousemove_brush_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_delay += 1
        time_mousemove_brush_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    ### mousemove count, later!!! detector per la mousedown prima.

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mousemove_count += 1
        time_special_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mousedown

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_count += 1
        time_mousedown_count += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_distance += 1
        time_mousedown_distance += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_arrival += 1
        time_mousedown_arrival += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_departure += 1
        time_mousedown_departure += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_airtime += 1
        time_mousedown_airtime += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_delay += 1
        time_mousedown_delay += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mouseup


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_distance += 1
        time_mouseup_distance += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_arrival += 1
        time_mouseup_arrival += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_departure += 1
        time_mouseup_departure += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_airtime += 1
        time_mouseup_airtime += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_delay += 1
        time_mouseup_delay += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### click

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "click") :
        
        count_click_count += 1
        time_click_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "click") :
        
        count_click_distance += 1
        time_click_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "click") :
        
        count_click_arrival += 1
        time_click_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "click") :
        
        count_click_departure += 1
        time_click_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "click") :
        
        count_click_airtime += 1
        time_click_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "click") :
        
        count_click_delay += 1
        time_click_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### dbclick

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_count += 1
        time_dbclick_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_distance += 1
        time_dbclick_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_arrival += 1
        time_dbclick_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_departure += 1
        time_dbclick_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_airtime += 1
        time_dbclick_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_delay += 1
        time_dbclick_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### wheel


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_distance += 1
        time_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_arrival += 1
        time_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_departure += 1
        time_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_airtime += 1
        time_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_delay += 1
        time_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### special wheel 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_wheel_distance += 1
        time_special_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_arrival += 1
        time_special_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_departure += 1
        time_special_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_airtime += 1
        time_special_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_delay += 1
        time_special_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### mouseout

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_count += 1
        time_mouseout_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_distance += 1
        time_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_arrival += 1
        time_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_departure += 1
        time_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)  
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_airtime += 1
        time_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_delay += 1
        time_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### special mouseout 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mouseout_distance += 1
        time_special_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
            
    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_arrival += 1
        time_special_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_departure += 1
        time_special_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_airtime += 1
        time_special_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_delay += 1
        time_special_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

#save on excel file##

if (count_mousemove_departure != 0) : sheet_total["A3"].value = (sheet_total["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_total["B3"].value = (sheet_total["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_total["C3"].value = (sheet_total["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_total["D3"].value = (sheet_total["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_total["E3"].value = (sheet_total["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_total["F3"].value = (sheet_total["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_total["G3"].value = (sheet_total["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_total["A6"].value = (sheet_total["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_total["B6"].value = (sheet_total["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_total["C6"].value = (sheet_total["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_total["D6"].value = (sheet_total["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_total["E6"].value = (sheet_total["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_total["F6"].value = (sheet_total["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_total["A10"].value = (sheet_total["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_total["B10"].value = (sheet_total["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_total["C10"].value = (sheet_total["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_total["D10"].value = (sheet_total["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_total["E10"].value = (sheet_total["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_total["A13"].value = (sheet_total["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_total["B13"].value = (sheet_total["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_total["C13"].value = (sheet_total["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_total["D13"].value = (sheet_total["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_total["E13"].value = (sheet_total["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_total["A17"].value = (sheet_total["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_total["B17"].value = (sheet_total["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_total["C17"].value = (sheet_total["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_total["D17"].value = (sheet_total["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_total["E17"].value = (sheet_total["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_total["F17"].value = (sheet_total["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_total["G17"].value = (sheet_total["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_total["A21"].value = (sheet_total["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_total["B21"].value = (sheet_total["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_total["C21"].value = (sheet_total["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_total["D21"].value = (sheet_total["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_total["E21"].value = (sheet_total["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_total["F21"].value = (sheet_total["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_total["A25"].value = (sheet_total["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_total["B25"].value = (sheet_total["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_total["C25"].value = (sheet_total["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_total["D25"].value = (sheet_total["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_total["E25"].value = (sheet_total["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_total["F25"].value = (sheet_total["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_total["A29"].value = (sheet_total["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_total["B29"].value = (sheet_total["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_total["C29"].value = (sheet_total["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_total["D29"].value = (sheet_total["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_total["E29"].value = (sheet_total["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_total["A33"].value = (sheet_total["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_total["B33"].value = (sheet_total["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_total["C33"].value = (sheet_total["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_total["D33"].value = (sheet_total["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_total["E33"].value = (sheet_total["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_total["A37"].value = (sheet_total["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_total["B37"].value = (sheet_total["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_total["C37"].value = (sheet_total["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_total["D37"].value = (sheet_total["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_total["E37"].value = (sheet_total["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_total["F37"].value = (sheet_total["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_total.save(filename="total_time.xlsx")


if (count_mousemove_departure != 0) : sheet_4["A3"].value = (sheet_4["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_4["B3"].value = (sheet_4["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_4["C3"].value = (sheet_4["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_4["D3"].value = (sheet_4["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_4["E3"].value = (sheet_4["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_4["F3"].value = (sheet_4["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_4["G3"].value = (sheet_4["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_4["A6"].value = (sheet_4["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_4["B6"].value = (sheet_4["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_4["C6"].value = (sheet_4["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_4["D6"].value = (sheet_4["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_4["E6"].value = (sheet_4["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_4["F6"].value = (sheet_4["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_4["A10"].value = (sheet_4["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_4["B10"].value = (sheet_4["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_4["C10"].value = (sheet_4["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_4["D10"].value = (sheet_4["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_4["E10"].value = (sheet_4["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_4["A13"].value = (sheet_4["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_4["B13"].value = (sheet_4["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_4["C13"].value = (sheet_4["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_4["D13"].value = (sheet_4["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_4["E13"].value = (sheet_4["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_4["A17"].value = (sheet_4["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_4["B17"].value = (sheet_4["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_4["C17"].value = (sheet_4["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_4["D17"].value = (sheet_4["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_4["E17"].value = (sheet_4["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_4["F17"].value = (sheet_4["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_4["G17"].value = (sheet_4["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_4["A21"].value = (sheet_4["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_4["B21"].value = (sheet_4["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_4["C21"].value = (sheet_4["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_4["D21"].value = (sheet_4["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_4["E21"].value = (sheet_4["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_4["F21"].value = (sheet_4["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_4["A25"].value = (sheet_4["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_4["B25"].value = (sheet_4["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_4["C25"].value = (sheet_4["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_4["D25"].value = (sheet_4["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_4["E25"].value = (sheet_4["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_4["F25"].value = (sheet_4["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_4["A29"].value = (sheet_4["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_4["B29"].value = (sheet_4["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_4["C29"].value = (sheet_4["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_4["D29"].value = (sheet_4["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_4["E29"].value = (sheet_4["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_4["A33"].value = (sheet_4["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_4["B33"].value = (sheet_4["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_4["C33"].value = (sheet_4["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_4["D33"].value = (sheet_4["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_4["E33"].value = (sheet_4["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_4["A37"].value = (sheet_4["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_4["B37"].value = (sheet_4["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_4["C37"].value = (sheet_4["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_4["D37"].value = (sheet_4["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_4["E37"].value = (sheet_4["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_4["F37"].value = (sheet_4["F37"].value + (time_dbclick_count/count_dbclick_count))/2


wb_4.save(filename="time_path_4.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

####also the time variables to 0!!!!

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#and the remember variables...

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0

###path 5!! da fare

for i in data["4"]:

    pathNumber = 4

    ### mousemove
    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_count += 1
        time_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_distance += 1
        time_mousemove_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_arrival += 1
        time_mousemove_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_departure += 1
        time_mousemove_departure += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_airtime += 1
        time_mousemove_airtime += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mousemove") :
        
        count_mousemove_delay += 1
        time_mousemove_delay += i[3]
        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

        ### mousemove brushing


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_distance += 1
        time_mousemove_brush_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_arrival += 1
        time_mousemove_brush_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_departure += 1
        time_mousemove_brush_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_airtime += 1
        time_mousemove_brush_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousemove") :
        
        count_mousemove_brush_delay += 1
        time_mousemove_brush_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    ### mousemove count, later!!! detector per la mousedown prima.

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mousemove")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mousemove_count += 1
        time_special_mousemove_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mousedown

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_count += 1
        time_mousedown_count += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_distance += 1
        time_mousedown_distance += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_arrival += 1
        time_mousedown_arrival += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_departure += 1
        time_mousedown_departure += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_airtime += 1
        time_mousedown_airtime += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mousedown") :
        
        count_mousedown_delay += 1
        time_mousedown_delay += i[3]
        remember_mousedown = 1
        remember_xpath_mousedown = i[0]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    
    ### mouseup


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_distance += 1
        time_mouseup_distance += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_arrival += 1
        time_mouseup_arrival += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_departure += 1
        time_mouseup_departure += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_airtime += 1
        time_mouseup_airtime += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "brush mouseup") :
        
        count_mouseup_delay += 1
        time_mouseup_delay += i[3]

        if(i[0] == remember_xpath_mousedown) and (remember_mousedown == 1):
            remember_xpath_mousedown = 0
            remember_mousedown = 0

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### click

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "click") :
        
        count_click_count += 1
        time_click_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "click") :
        
        count_click_distance += 1
        time_click_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "click") :
        
        count_click_arrival += 1
        time_click_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "click") :
        
        count_click_departure += 1
        time_click_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "click") :
        
        count_click_airtime += 1
        time_click_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "click") :
        
        count_click_delay += 1
        time_click_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)


    ### dbclick

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_count += 1
        time_dbclick_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_distance += 1
        time_dbclick_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_arrival += 1
        time_dbclick_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_departure += 1
        time_dbclick_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_airtime += 1
        time_dbclick_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "dbclick") :
        
        count_dbclick_delay += 1
        time_dbclick_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### wheel


    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_distance += 1
        time_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_arrival += 1
        time_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_departure += 1
        time_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_airtime += 1
        time_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        count_wheel_delay += 1
        time_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### special wheel 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_wheel_distance += 1
        time_special_wheel_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_arrival += 1
        time_special_wheel_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_departure += 1
        time_special_wheel_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_airtime += 1
        time_special_wheel_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_wheel_delay += 1
        time_special_wheel_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
    

    ### mouseout

    if(i[0] == "/html[1]/body[1]/div[2]/div[1]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_count += 1
        time_mouseout_count += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_distance += 1
        time_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_arrival += 1
        time_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_departure += 1
        time_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)  
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_airtime += 1
        time_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "mouseout") :
        
        count_mouseout_delay += 1
        time_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    ### special mouseout 

    if(i[0] == "/html[1]/body[1]/div[2]/div[2]/canvas[1]") and (i[1] == "wheel")  and (remember_mousedown == 1) and (remember_xpath_mousedown == i[0]):
        
        special_mouseout_distance += 1
        time_special_mouseout_distance += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)
            
    if(i[0] == "/html[1]/body[1]/div[2]/div[3]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_arrival += 1
        time_special_mouseout_arrival += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[4]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_departure += 1
        time_special_mouseout_departure += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[6]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_airtime += 1
        time_special_mouseout_airtime += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)   
            violationsFound[pathNumber].append(toWrite)

    if(i[0] == "/html[1]/body[1]/div[2]/div[7]/canvas[1]") and (i[1] == "wheel") :
        
        special_mouseout_delay += 1
        time_special_mouseout_delay += i[3]

        if(i[4] != 0) and (i[4] != -1): 
            toWrite = "violation of level " + str(i[4]) + " on " + str(i[0]) + str(i[1]) + " with time: " + str(i[3]) + " with details: " + str(i[2]) + " in path: " + str(pathNumber)
            violationsFound[pathNumber].append(toWrite)

# save on excel

if (count_mousemove_departure != 0) : sheet_total["A3"].value = (sheet_total["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_total["B3"].value = (sheet_total["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_total["C3"].value = (sheet_total["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_total["D3"].value = (sheet_total["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_total["E3"].value = (sheet_total["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_total["F3"].value = (sheet_total["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_total["G3"].value = (sheet_total["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_total["A6"].value = (sheet_total["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_total["B6"].value = (sheet_total["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_total["C6"].value = (sheet_total["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_total["D6"].value = (sheet_total["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_total["E6"].value = (sheet_total["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_total["F6"].value = (sheet_total["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_total["A10"].value = (sheet_total["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_total["B10"].value = (sheet_total["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_total["C10"].value = (sheet_total["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_total["D10"].value = (sheet_total["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_total["E10"].value = (sheet_total["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_total["A13"].value = (sheet_total["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_total["B13"].value = (sheet_total["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_total["C13"].value = (sheet_total["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_total["D13"].value = (sheet_total["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_total["E13"].value = (sheet_total["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_total["A17"].value = (sheet_total["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_total["B17"].value = (sheet_total["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_total["C17"].value = (sheet_total["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_total["D17"].value = (sheet_total["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_total["E17"].value = (sheet_total["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_total["F17"].value = (sheet_total["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_total["G17"].value = (sheet_total["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_total["A21"].value = (sheet_total["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_total["B21"].value = (sheet_total["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_total["C21"].value = (sheet_total["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_total["D21"].value = (sheet_total["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_total["E21"].value = (sheet_total["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_total["F21"].value = (sheet_total["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_total["A25"].value = (sheet_total["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_total["B25"].value = (sheet_total["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_total["C25"].value = (sheet_total["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_total["D25"].value = (sheet_total["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_total["E25"].value = (sheet_total["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_total["F25"].value = (sheet_total["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_total["A29"].value = (sheet_total["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_total["B29"].value = (sheet_total["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_total["C29"].value = (sheet_total["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_total["D29"].value = (sheet_total["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_total["E29"].value = (sheet_total["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_total["A33"].value = (sheet_total["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_total["B33"].value = (sheet_total["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_total["C33"].value = (sheet_total["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_total["D33"].value = (sheet_total["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_total["E33"].value = (sheet_total["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_total["A37"].value = (sheet_total["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_total["B37"].value = (sheet_total["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_total["C37"].value = (sheet_total["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_total["D37"].value = (sheet_total["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_total["E37"].value = (sheet_total["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_total["F37"].value = (sheet_total["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_total.save(filename="total_time.xlsx")



if (count_mousemove_departure != 0) : sheet_5["A3"].value = (sheet_5["A3"].value + (time_mousemove_departure/count_mousemove_departure))/2
if (count_mousemove_distance != 0) : sheet_5["B3"].value = (sheet_5["B3"].value + (time_mousemove_distance/count_mousemove_distance))/2
if (count_mousemove_delay != 0) : sheet_5["C3"].value = (sheet_5["C3"].value + (time_mousemove_delay/count_mousemove_delay))/2
if (count_mousemove_airtime != 0) : sheet_5["D3"].value = (sheet_5["D3"].value + (time_mousemove_airtime/count_mousemove_airtime))/2
if (count_mousemove_arrival != 0) : sheet_5["E3"].value = (sheet_5["E3"].value + (time_mousemove_arrival/count_mousemove_arrival))/2
if (count_mousemove_count != 0) : sheet_5["F3"].value = (sheet_5["F3"].value + (time_mousemove_count/count_mousemove_count))/2
if (special_mousemove_count != 0) : sheet_5["G3"].value = (sheet_5["G3"].value + (time_special_mousemove_count/special_mousemove_count))/2

if (count_mousedown_departure != 0) : sheet_5["A6"].value = (sheet_5["A6"].value + (time_mousedown_departure/count_mousedown_departure))/2
if (count_mousedown_distance != 0) : sheet_5["B6"].value = (sheet_5["B6"].value + (time_mousedown_distance/count_mousedown_distance))/2
if (count_mousedown_delay != 0) : sheet_5["C6"].value = (sheet_5["C6"].value + (time_mousedown_delay/count_mousedown_delay))/2
if (count_mousedown_airtime != 0) : sheet_5["D6"].value = (sheet_5["D6"].value + (time_mousedown_airtime/count_mousedown_airtime))/2
if (count_mousedown_arrival != 0) : sheet_5["E6"].value = (sheet_5["E6"].value + (time_mousedown_arrival/count_mousedown_arrival))/2
if (count_mousedown_count != 0) : sheet_5["F6"].value = (sheet_5["F6"].value + (time_mousedown_count/count_mousedown_count))/2

if (count_mousemove_brush_departure != 0) : sheet_5["A10"].value = (sheet_5["A10"].value + (time_mousemove_brush_departure/count_mousemove_brush_departure))/2
if (count_mousemove_brush_distance != 0) : sheet_5["B10"].value = (sheet_5["B10"].value + (time_mousemove_brush_distance/count_mousemove_brush_distance))/2
if (count_mousemove_brush_delay != 0) : sheet_5["C10"].value = (sheet_5["C10"].value + (time_mousemove_brush_delay/count_mousemove_brush_delay))/2
if (count_mousemove_brush_airtime != 0) : sheet_5["D10"].value = (sheet_5["D10"].value + (time_mousemove_brush_airtime/count_mousemove_brush_airtime))/2
if (count_mousemove_brush_arrival != 0) : sheet_5["E10"].value = (sheet_5["E10"].value + (time_mousemove_brush_arrival/count_mousemove_brush_arrival))/2

if (count_mouseup_departure != 0) : sheet_5["A13"].value = (sheet_5["A13"].value + (time_mouseup_departure/count_mouseup_departure))/2
if (count_mouseup_distance != 0) : sheet_5["B13"].value = (sheet_5["B13"].value + (time_mouseup_distance/count_mouseup_distance))/2
if (count_mouseup_delay != 0) : sheet_5["C13"].value = (sheet_5["C13"].value + (time_mouseup_delay/count_mouseup_delay))/2
if (count_mouseup_airtime != 0) : sheet_5["D13"].value = (sheet_5["D13"].value + (time_mouseup_airtime/count_mouseup_airtime))/2
if (count_mouseup_arrival != 0) : sheet_5["E13"].value = (sheet_5["E13"].value + (time_mouseup_arrival/count_mouseup_arrival))/2

if (count_click_departure != 0) : sheet_5["A17"].value = (sheet_5["A17"].value + (time_click_departure/count_click_departure))/2
if (count_click_distance != 0) : sheet_5["B17"].value = (sheet_5["B17"].value + (time_click_distance/count_click_distance))/2
if (count_click_delay != 0) : sheet_5["C17"].value = (sheet_5["C17"].value + (time_click_delay/count_click_delay))/2
if (count_click_airtime != 0) : sheet_5["D17"].value = (sheet_5["D17"].value + (time_click_airtime/count_click_airtime))/2
if (count_click_arrival != 0) : sheet_5["E17"].value = (sheet_5["E17"].value + (time_click_arrival/count_click_arrival))/2
if (count_click_count != 0) : sheet_5["F17"].value = (sheet_5["F17"].value + (time_click_count/count_click_count))/2
if (special_click_count != 0) : sheet_5["G17"].value = (sheet_5["G17"].value + (time_special_click_count/special_click_count))/2
###
if (count_mouseout_departure != 0) : sheet_5["A21"].value = (sheet_5["A21"].value + (time_mouseout_departure/count_mouseout_departure))/2
if (count_mouseout_distance != 0) : sheet_5["B21"].value = (sheet_5["B21"].value + (time_mouseout_distance/count_mouseout_distance))/2
if (count_mouseout_delay != 0) : sheet_5["C21"].value = (sheet_5["C21"].value + (time_mouseout_delay/count_mouseout_delay))/2
if (count_mouseout_airtime != 0) : sheet_5["D21"].value = (sheet_5["D21"].value + (time_mouseout_airtime/count_mouseout_airtime))/2
if (count_mouseout_arrival != 0) : sheet_5["E21"].value = (sheet_5["E21"].value + (time_mouseout_arrival/count_mouseout_arrival))/2
if (count_mouseout_count != 0) : sheet_5["F21"].value = (sheet_5["F21"].value + (time_mouseout_count/count_mouseout_count))/2

if (special_mouseout_departure != 0) : sheet_5["A25"].value = (sheet_5["A25"].value + (time_special_mouseout_departure/special_mouseout_departure))/2
if (special_mouseout_distance != 0) : sheet_5["B25"].value = (sheet_5["B25"].value + (time_special_mouseout_distance/special_mouseout_distance))/2
if (special_mouseout_delay != 0) : sheet_5["C25"].value = (sheet_5["C25"].value + (time_special_mouseout_delay/special_mouseout_delay))/2
if (special_mouseout_airtime != 0) : sheet_5["D25"].value = (sheet_5["D25"].value + (time_special_mouseout_airtime/special_mouseout_airtime))/2
if (special_mouseout_arrival != 0) : sheet_5["E25"].value = (sheet_5["E25"].value + (time_special_mouseout_arrival/special_mouseout_arrival))/2
if (special_mouseout_count != 0) : sheet_5["F25"].value = (sheet_5["F25"].value + (time_special_mouseout_count/special_mouseout_count))/2

if (count_wheel_departure != 0) : sheet_5["A29"].value = (sheet_5["A29"].value + (time_wheel_departure/count_wheel_departure))/2
if (count_wheel_distance != 0) : sheet_5["B29"].value = (sheet_5["B29"].value + (time_wheel_distance/count_wheel_distance))/2
if (count_wheel_delay != 0) : sheet_5["C29"].value = (sheet_5["C29"].value + (time_wheel_delay/count_wheel_delay))/2
if (count_wheel_airtime != 0) : sheet_5["D29"].value = (sheet_5["D29"].value + (time_wheel_airtime/count_wheel_airtime))/2
if (count_wheel_arrival != 0) : sheet_5["E29"].value = (sheet_5["E29"].value + (time_wheel_arrival/count_wheel_arrival))/2

if (special_wheel_departure != 0) : sheet_5["A33"].value = (sheet_5["A33"].value + (time_special_wheel_departure/special_wheel_departure))/2
if (special_wheel_distance != 0) : sheet_5["B33"].value = (sheet_5["B33"].value + (time_special_wheel_distance/special_wheel_distance))/2
if (special_wheel_delay != 0) : sheet_5["C33"].value = (sheet_5["C33"].value + (time_special_wheel_delay/special_wheel_delay))/2
if (special_wheel_airtime != 0) : sheet_5["D33"].value = (sheet_5["D33"].value + (time_special_wheel_airtime/special_wheel_airtime))/2
if (special_wheel_arrival != 0) : sheet_5["E33"].value = (sheet_5["E33"].value + (time_special_wheel_arrival/special_wheel_arrival))/2

if (count_dbclick_departure != 0) : sheet_5["A37"].value = (sheet_5["A37"].value + (time_dbclick_departure/count_dbclick_departure))/2
if (count_dbclick_distance != 0) : sheet_5["B37"].value = (sheet_5["B37"].value + (time_dbclick_distance/count_dbclick_distance))/2
if (count_dbclick_delay != 0) : sheet_5["C37"].value = (sheet_5["C37"].value + (time_dbclick_delay/count_dbclick_delay))/2
if (count_dbclick_airtime != 0) : sheet_5["D37"].value = (sheet_5["D37"].value + (time_dbclick_airtime/count_dbclick_airtime))/2
if (count_dbclick_arrival != 0) : sheet_5["E37"].value = (sheet_5["E37"].value + (time_dbclick_arrival/count_dbclick_arrival))/2
if (count_dbclick_count != 0) : sheet_5["F37"].value = (sheet_5["F37"].value + (time_dbclick_count/count_dbclick_count))/2

wb_5.save(filename="time_path_5.xlsx")

count_mousemove_delay = 0
count_mousemove_airtime = 0
count_mousemove_distance = 0
count_mousemove_arrival = 0
count_mousemove_departure = 0
count_mousemove_count = 0

special_mousemove_count = 0


#mousedown

count_mousedown_delay = 0
count_mousedown_airtime = 0
count_mousedown_distance = 0
count_mousedown_arrival = 0
count_mousedown_departure = 0
count_mousedown_count = 0

count_mousemove_brush_delay = 0
count_mousemove_brush_airtime = 0
count_mousemove_brush_distance = 0
count_mousemove_brush_arrival = 0
count_mousemove_brush_departure = 0

#mouseup

count_mouseup_delay = 0
count_mouseup_airtime = 0
count_mouseup_distance = 0
count_mouseup_arrival = 0
count_mouseup_departure = 0

#click

count_click_delay = 0
count_click_airtime = 0
count_click_distance = 0
count_click_arrival = 0
count_click_departure = 0
count_click_count = 0

special_click_count = 0

#dbclick

count_dbclick_delay = 0
count_dbclick_airtime = 0
count_dbclick_distance = 0
count_dbclick_arrival = 0
count_dbclick_departure = 0
count_dbclick_count = 0

#wheel

count_wheel_delay = 0
count_wheel_airtime = 0
count_wheel_distance = 0
count_wheel_arrival = 0
count_wheel_departure = 0

special_wheel_delay = 0
special_wheel_airtime = 0
special_wheel_distance = 0
special_wheel_arrival = 0
special_wheel_departure = 0

#mouseout

count_mouseout_delay = 0
count_mouseout_airtime = 0
count_mouseout_distance = 0
count_mouseout_arrival = 0
count_mouseout_departure = 0
count_mouseout_count = 0
special_mouseout_count = 0

special_mouseout_delay = 0
special_mouseout_airtime = 0
special_mouseout_distance = 0
special_mouseout_arrival = 0
special_mouseout_departure = 0

####also the time variables to 0!!!!

#mousemove
time_mousemove_delay = 0
time_mousemove_airtime = 0
time_mousemove_distance = 0
time_mousemove_arrival = 0
time_mousemove_departure = 0
time_mousemove_count = 0

time_special_mousemove_count = 0


#mousedown

time_mousedown_delay = 0
time_mousedown_airtime = 0
time_mousedown_distance = 0
time_mousedown_arrival = 0
time_mousedown_departure = 0
time_mousedown_count = 0

time_mousemove_brush_delay = 0
time_mousemove_brush_airtime = 0
time_mousemove_brush_distance = 0
time_mousemove_brush_arrival = 0
time_mousemove_brush_departure = 0

#mouseup

time_mouseup_delay = 0
time_mouseup_airtime = 0
time_mouseup_distance = 0
time_mouseup_arrival = 0
time_mouseup_departure = 0

#click

time_click_delay = 0
time_click_airtime = 0
time_click_distance = 0
time_click_arrival = 0
time_click_departure = 0
time_click_count = 0

time_special_click_count = 0

#dbclick

time_dbclick_delay = 0
time_dbclick_airtime = 0
time_dbclick_distance = 0
time_dbclick_arrival = 0
time_dbclick_departure = 0
time_dbclick_count = 0

#wheel

time_wheel_delay = 0
time_wheel_airtime = 0
time_wheel_distance = 0
time_wheel_arrival = 0
time_wheel_departure = 0

time_special_wheel_delay = 0
time_special_wheel_airtime = 0
time_special_wheel_distance = 0
time_special_wheel_arrival = 0
time_special_wheel_departure = 0

#mouseout

time_mouseout_delay = 0
time_mouseout_airtime = 0
time_mouseout_distance = 0
time_mouseout_arrival = 0
time_mouseout_departure = 0
time_mouseout_count = 0
time_special_mouseout_count = 0

time_special_mouseout_delay = 0
time_special_mouseout_airtime = 0
time_special_mouseout_distance = 0
time_special_mouseout_arrival = 0
time_special_mouseout_departure = 0

#and the remember variables...

remember = 0
remember_mousedown = 0
remember_xpath_mousedown = 0
    
        
    
  

# Closing file
with open('violationsFound_' + "falcon_7M" + '_50' + '.json', 'w') as fp:
            json.dump(violationsFound, fp,  indent=4)

f.close()